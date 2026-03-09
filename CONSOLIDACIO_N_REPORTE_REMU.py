import streamlit as st
import io
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter


# ----------------------------
# Utilidades encabezados
# ----------------------------

def norm(x):
    if x is None:
        return ""
    return str(x).strip()

def find_header_rows(ws, scan_rows=80):
    for r in range(1, scan_rows + 1):
        v1 = norm(ws.cell(row=r, column=1).value)
        if v1.upper() == "PROCESO":
            if r < 3:
                raise ValueError(f"Header3 en fila {r}, no hay 2 filas arriba.")
            return (r - 2, r - 1, r)

    for r in range(1, scan_rows + 1):
        row_vals = [norm(ws.cell(row=r, column=c).value).upper() for c in range(1, 20)]
        if "PROCESO" in row_vals and "FOLIO" in row_vals:
            if r < 3:
                raise ValueError(f"Header3 en fila {r}, no hay 2 filas arriba.")
            return (r - 2, r - 1, r)

    raise ValueError("No se detectó el header3 (fila con 'PROCESO').")


def get_max_col_smart(ws, header3_row, scan_extra=50):
    max_col = ws.max_column
    last_nonempty = 0
    for c in range(1, max_col + 1):
        if norm(ws.cell(row=header3_row, column=c).value) != "":
            last_nonempty = c
    return max(last_nonempty + scan_extra, last_nonempty, 1)


def build_sheet_header_keys(ws, h1, h2, h3):
    max_col = get_max_col_smart(ws, h3)
    keys = []
    for c in range(1, max_col + 1):
        a = norm(ws.cell(row=h1, column=c).value)
        b = norm(ws.cell(row=h2, column=c).value)
        d = norm(ws.cell(row=h3, column=c).value)
        if a == "" and b == "" and d == "":
            keys.append((f"__BLANK__{get_column_letter(c)}", "", ""))
        else:
            keys.append((a, b, d))
    return keys


def consolidate_by_header_triplet(file_bytes, sheet_names, add_source_col=True):
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=False, keep_links=False)

    global_cols = []
    global_index = {}
    per_sheet = {}

    for sname in sheet_names:
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        h1, h2, h3 = find_header_rows(ws)
        keys = build_sheet_header_keys(ws, h1, h2, h3)
        max_col = len(keys)
        colmap = {}
        for c, k in enumerate(keys, start=1):
            if k not in global_index:
                global_index[k] = len(global_cols)
                global_cols.append(k)
            colmap[c] = global_index[k]
        per_sheet[sname] = {
            "h1": h1, "h2": h2, "h3": h3,
            "start_data_row": h3 + 1,
            "max_col": max_col,
            "colmap": colmap
        }

    if not global_cols:
        raise ValueError("No se pudo construir columnas globales.")

    wb_out = Workbook(write_only=True)
    ws_out = wb_out.create_sheet("CONSOLIDADO")

    offset = 1 if add_source_col else 0
    row_h1 = (["ORIGEN_HOJA"] if add_source_col else []) + [k[0] for k in global_cols]
    row_h2 = ([""] if add_source_col else []) + [k[1] for k in global_cols]
    row_h3 = ([""] if add_source_col else []) + [k[2] for k in global_cols]
    ws_out.append(row_h1)
    ws_out.append(row_h2)
    ws_out.append(row_h3)

    for sname in sheet_names:
        if sname not in per_sheet:
            continue
        ws = wb[sname]
        info = per_sheet[sname]
        for row in ws.iter_rows(min_row=info["start_data_row"], max_col=info["max_col"], values_only=True):
            if row is None:
                continue
            if all((v is None or v == "") for v in row):
                continue
            out = [None] * (len(global_cols) + offset)
            if add_source_col:
                out[0] = sname
            for c, v in enumerate(row, start=1):
                gi = info["colmap"].get(c)
                if gi is None:
                    continue
                out[gi + offset] = v
            ws_out.append(out)

    output = io.BytesIO()
    wb_out.save(output)
    output.seek(0)
    return output


# ----------------------------
# Interfaz Streamlit
# ----------------------------

st.title("📊 Consolidación de Reportes")
st.markdown("Consolida múltiples hojas de un Excel con encabezado triple (3 filas).")

uploaded_file = st.file_uploader("Selecciona archivo Excel de entrada", type=["xlsx"])

if uploaded_file:
    file_bytes = uploaded_file.read()
    wb_preview = load_workbook(io.BytesIO(file_bytes), read_only=True)
    sheet_names = wb_preview.sheetnames

    st.markdown("**Selecciona las hojas a consolidar (en orden):**")
    selected_sheets = st.multiselect("Hojas disponibles", sheet_names, default=sheet_names[:3])

    add_source = st.checkbox("Agregar columna ORIGEN_HOJA (recomendado)", value=True)

    if st.button("CONSOLIDAR", type="primary"):
        if not selected_sheets:
            st.warning("Selecciona al menos una hoja.")
        else:
            with st.spinner("Consolidando..."):
                try:
                    output = consolidate_by_header_triplet(file_bytes, selected_sheets, add_source)
                    st.success("✅ Consolidado generado correctamente.")
                    st.download_button(
                        label="📥 Descargar archivo consolidado",
                        data=output,
                        file_name=uploaded_file.name.replace(".xlsx", "_CONSOLIDADO.xlsx"),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                except Exception as e:
                    st.error(f"Error al consolidar: {e}")
