# -*- coding: utf-8 -*-
"""
REDISTRIBUCIÓN PRAPS vs D.A.P (ADMINISTRACION) — UI + Motor (Streamlit)
======================================================================

✅ CONSOLIDADO: encabezados compuestos por 3 filas (se concatenan A|B|C por columna)
✅ BLOQUE DESCUENTO: tabla transpuesta con columnas TIPO DE HABER, CODIGO, DESCRIPCIÓN
✅ Limpieza Reintegros: borra pares (+/-) por RUT-DV, MES PAGO, abs(TOTAL HABER)
✅ Redistribución trazable:
   - Fila ORIGEN: MONTO_FINANCIADO = HABER_NETO
   - Fila DAP (S/R + DAP): TOTAL HABER = 0, MONTO_FINANCIADO = suma descuentos bloque
     + detalle por concepto (columnas del bloque en positivo)
✅ Salida Excel:
   1) CONSOLIDADO_REDISTRIBUIDO
   2) RESUMEN_CC_PROGRAMA
   3) ESTADO_RESULTADOS
   4) AUDITORIA
   + Gráficos en ESTADO_RESULTADOS (openpyxl)

USO (UI):
  streamlit run app_redistribucion.py

USO (CLI opcional):
  python app_redistribucion.py --input "archivo.xlsx" --output "salida.xlsx"
"""

from __future__ import annotations

import argparse
import io
import os
import re
import sys
import uuid
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList


# -------------------------
# Detección robusta Streamlit
# -------------------------
def running_under_streamlit() -> bool:
    """
    Detecta ejecución bajo "streamlit run" de forma robusta.
    No depende de sys.argv (que en el script normalmente NO trae --server.*).
    """
    try:
        from streamlit.runtime.scriptrunner import get_script_run_ctx  # type: ignore
        return get_script_run_ctx() is not None
    except Exception:
        return False


# =========================
# Utils: normalización
# =========================
def _norm(x: object) -> str:
    if x is None:
        return ""
    s = str(x).replace("\u00A0", " ")
    s = s.strip()
    s = re.sub(r"\s+", " ", s)
    return s.upper()


def _join3(a: object, b: object, c: object) -> str:
    return f"{_norm(a)} | {_norm(b)} | {_norm(c)}"


def _safe_to_numeric(s: pd.Series) -> pd.Series:
    return pd.to_numeric(s, errors="coerce").fillna(0)


def _contains_ci(series: pd.Series, needle: str) -> pd.Series:
    return series.astype(str).str.contains(needle, case=False, na=False)




# =========================
# Regla: descuentos que SÍ rebajan el neto pero NO se redistribuyen a DAP
# (financiamiento externo / no imputable a CC-Programa DAP)
# =========================
def _is_discount_no_redistrib(colname: str) -> bool:
    """Detecta columnas del tipo 'TIPO | CODIGO | DESCRIPCIÓN' que deben
    descontar del neto pero NO generar redistribución a DAP.
    Regla actual: HABER NO PERMANENTE + código 63 o 64 (independiente de la glosa).
    """
    base = str(colname).split("__DUP")[0]
    parts = [p.strip() for p in base.split("|")]
    if len(parts) < 2:
        return False
    tipo = _norm(parts[0])
    # tolera variaciones: PERMANENTE / PERMANANTE, espacios, etc.
    if not tipo.startswith("HABER NO PERMAN"):
        return False
    codigo_raw = parts[1]
    codigo = re.sub(r"\D+", "", str(codigo_raw))
    return codigo in {"63", "64"}
# =========================
# Auditoría
# =========================
@dataclass
class AuditReport:
    missing_discount_columns: List[str]
    duplicate_discount_keys_in_block: List[str]
    reintegro_pairs_deleted: int
    reintegro_unpaired_rows: int
    reintegro_unpaired_sample: pd.DataFrame
    neto_negative_rows: int
    neto_negative_sample: pd.DataFrame


# =========================
# Lectura encabezado compuesto (3 filas)
# =========================
def read_consolidado_with_compound_header(
    file_like,
    sheet_name: str,
    header_rows: int = 3,
) -> Tuple[pd.DataFrame, List[str], List[Tuple[str, str, str]]]:
    header_raw = pd.read_excel(
        file_like, sheet_name=sheet_name, header=None, nrows=header_rows, engine="openpyxl"
    )
    data_raw = pd.read_excel(
        file_like, sheet_name=sheet_name, header=None, skiprows=header_rows, engine="openpyxl"
    )

    ncols = max(header_raw.shape[1], data_raw.shape[1])
    if header_raw.shape[1] < ncols:
        for _ in range(ncols - header_raw.shape[1]):
            header_raw[header_raw.shape[1]] = ""
    if data_raw.shape[1] < ncols:
        for _ in range(ncols - data_raw.shape[1]):
            data_raw[data_raw.shape[1]] = np.nan

    h1 = header_raw.iloc[0].tolist()
    h2 = header_raw.iloc[1].tolist() if header_rows >= 2 else [""] * ncols
    h3 = header_raw.iloc[2].tolist() if header_rows >= 3 else [""] * ncols

    parts: List[Tuple[str, str, str]] = []
    colnames: List[str] = []
    for i in range(ncols):
        p1 = "" if h1[i] is None else str(h1[i])
        p2 = "" if h2[i] is None else str(h2[i])
        p3 = "" if h3[i] is None else str(h3[i])

        key = _join3(p1, p2, p3)
        if key.strip(" |") == "":
            key = f"UNNAMED_{i}"
        colnames.append(key)
        parts.append((_norm(p1), _norm(p2), _norm(p3)))

    # Evita duplicados de columnas
    seen: Dict[str, int] = {}
    unique_cols = []
    for c in colnames:
        if c not in seen:
            seen[c] = 0
            unique_cols.append(c)
        else:
            seen[c] += 1
            unique_cols.append(f"{c}__DUP{seen[c]}")
    colnames = unique_cols

    df = data_raw.copy()
    df.columns = colnames
    df = df.dropna(how="all")

    return df, colnames, parts


# =========================
# Bloque descuento -> llaves
# =========================
def read_discount_block_keys(
    file_like,
    sheet_name: str = "BLOQUE DESCUENTO",
    tipo_col: str = "TIPO DE HABER",
    codigo_col: str = "CODIGO",
    desc_col: str = "DESCRIPCIÓN",
) -> Tuple[List[str], List[str]]:
    block = pd.read_excel(file_like, sheet_name=sheet_name, engine="openpyxl")
    block.columns = [c.strip() for c in block.columns.astype(str)]

    def find_col(candidates: List[str]) -> str:
        for cand in candidates:
            for col in block.columns:
                if _norm(col) == _norm(cand):
                    return col
        raise KeyError(
            f"No encontré columna en BLOQUE DESCUENTO: {candidates}. Disponibles: {list(block.columns)}"
        )

    tipo_c = find_col([tipo_col, "TIPO DE HABER", "TIPO DE HABER "])
    cod_c = find_col([codigo_col, "CODIGO", "CÓDIGO"])
    des_c = find_col([desc_col, "DESCRIPCIÓN", "DESCRIPCION", "GLOSA"])

    keys_raw = block[[tipo_c, cod_c, des_c]].copy().dropna(how="all")
    keys = keys_raw.apply(lambda r: _join3(r.iloc[0], r.iloc[1], r.iloc[2]), axis=1).tolist()

    s = pd.Series(keys)
    dup_keys = s[s.duplicated(keep=False)].unique().tolist()

    seen = set()
    keys_dedup = []
    for k in keys:
        if k not in seen:
            seen.add(k)
            keys_dedup.append(k)

    return keys_dedup, dup_keys


# =========================
# Encontrar columnas clave por tokens
# =========================
def find_column(colnames: List[str], must_contain_any: List[str]) -> str:
    targets = [_norm(x) for x in must_contain_any]
    for c in colnames:
        cN = _norm(c)
        for t in targets:
            if t in cN:
                return c
    raise KeyError(f"No pude localizar columna con tokens {must_contain_any}. Revisa encabezados compuestos.")


# =========================
# Reintegros: borrar pares (+/-)
# =========================
def remove_reintegro_pairs(
    df: pd.DataFrame,
    col_proceso: str,
    col_rut: str,
    col_mes: str,
    col_total_haber: str,
    tolerance_pesos: int = 0,
) -> Tuple[pd.DataFrame, int, pd.DataFrame]:
    work = df.copy()
    work["_TOTAL_NUM"] = _safe_to_numeric(work[col_total_haber])

    is_reintegro = _contains_ci(work[col_proceso], "REINTEG")
    rein_df = work[is_reintegro & (work["_TOTAL_NUM"] != 0)].copy()
    if rein_df.empty:
        return df, 0, pd.DataFrame()

    abs_val = rein_df["_TOTAL_NUM"].abs()
    if tolerance_pesos > 0:
        abs_key = (abs_val / tolerance_pesos).round().astype("Int64") * tolerance_pesos
    else:
        abs_key = abs_val.round(0).astype("Int64")

    rein_df["_KEY"] = (
        rein_df[col_rut].astype(str).fillna("")
        + "||"
        + rein_df[col_mes].astype(str).fillna("")
        + "||"
        + abs_key.astype(str).fillna("")
    )

    cand = work.copy()
    cand["_TOTAL_NUM"] = _safe_to_numeric(cand[col_total_haber])
    abs_val_c = cand["_TOTAL_NUM"].abs()
    if tolerance_pesos > 0:
        abs_key_c = (abs_val_c / tolerance_pesos).round().astype("Int64") * tolerance_pesos
    else:
        abs_key_c = abs_val_c.round(0).astype("Int64")

    cand["_KEY"] = (
        cand[col_rut].astype(str).fillna("")
        + "||"
        + cand[col_mes].astype(str).fillna("")
        + "||"
        + abs_key_c.astype(str).fillna("")
    )

    to_delete = set()
    pairs_deleted = 0

    keys_interest = set(rein_df["_KEY"].unique().tolist())
    subset = cand[cand["_KEY"].isin(keys_interest)]

    for key, g in subset.groupby("_KEY", sort=False):
        pos_idx = g[g["_TOTAL_NUM"] > 0].index.tolist()
        neg_idx = g[g["_TOTAL_NUM"] < 0].index.tolist()
        n = min(len(pos_idx), len(neg_idx))
        if n > 0:
            for i in range(n):
                to_delete.add(pos_idx[i])
                to_delete.add(neg_idx[i])
            pairs_deleted += n

    deleted_mask = df.index.isin(list(to_delete))
    df_clean = df.loc[~deleted_mask].copy()
    rein_unpaired = df.loc[is_reintegro & ~deleted_mask].copy()

    return df_clean, pairs_deleted, rein_unpaired


# =========================
# Motor principal
# =========================
def generate_redistribution(
    input_path_or_buffer,
    sheet_consolidado: str,
    sheet_bloque: str = "BLOQUE DESCUENTO",
    header_rows: int = 3,
    cc_dap: str = "S/R",
    prog_dap: str = "D.A.P. (ADMINISTRACION)",
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, AuditReport]:

    file_like = input_path_or_buffer

    # 1) Leer CONSOLIDADO con header compuesto
    df, colnames, _parts = read_consolidado_with_compound_header(
        file_like, sheet_name=sheet_consolidado, header_rows=header_rows
    )

    # 2) Columnas clave
    col_proceso = find_column(colnames, ["PROCESO"])
    col_mes = find_column(colnames, ["MES PAGO", "MES", "PAGO MES"])
    col_rut = find_column(colnames, ["RUT-DV", "RUT DV", "RUT"])
    col_cc = find_column(colnames, ["CENTRO DE COSTO"])
    col_prog = find_column(colnames, ["PROGRAMA"])
    col_total_haber = find_column(colnames, ["TOTAL HABER"])

    # 3) Bloque descuentos: llaves
    block_keys, dup_block_keys = read_discount_block_keys(file_like, sheet_name=sheet_bloque)

    # 4) Match exacto de llaves vs header compuesto
    norm_to_actual: Dict[str, str] = {}
    for c in colnames:
        base = c.split("__DUP")[0]
        norm_to_actual[_norm(base)] = c

    discount_cols: List[str] = []
    missing_cols: List[str] = []
    for k in block_keys:
        nk = _norm(k)
        if nk in norm_to_actual:
            discount_cols.append(norm_to_actual[nk])
        else:
            missing_cols.append(k)


    # Separar: descuentos totales vs descuentos redistribuibles a DAP
    discount_cols_all: List[str] = discount_cols
    discount_cols_no_redistrib: List[str] = [c for c in discount_cols_all if _is_discount_no_redistrib(c)]
    discount_cols_redistrib: List[str] = [c for c in discount_cols_all if c not in set(discount_cols_no_redistrib)]

    # 5) Limpieza reintegros
    df_clean, rein_pairs_deleted, rein_unpaired = remove_reintegro_pairs(
        df=df,
        col_proceso=col_proceso,
        col_rut=col_rut,
        col_mes=col_mes,
        col_total_haber=col_total_haber,
        tolerance_pesos=0,
    )

    # 6) Cálculos
    work = df_clean.copy()

    total_haber_num = _safe_to_numeric(work[col_total_haber])
    if discount_cols_all:
        for c in discount_cols_all:
            work[c] = _safe_to_numeric(work[c])

    work["TOTAL_HABER_BRUTO"] = total_haber_num
    work["DESCUENTOS_BLOQUE"] = work[discount_cols_all].sum(axis=1) if discount_cols_all else 0.0
    work["DESCUENTOS_REDISTRIBUIBLES"] = work[discount_cols_redistrib].sum(axis=1) if discount_cols_redistrib else 0.0
    work["DESCUENTOS_NO_REDISTRIB"] = work[discount_cols_no_redistrib].sum(axis=1) if discount_cols_no_redistrib else 0.0
    work["HABER_NETO"] = work["TOTAL_HABER_BRUTO"] - work["DESCUENTOS_BLOQUE"]

    # Trazabilidad
    work["ID_ORIGEN_FILA"] = [str(uuid.uuid4()) for _ in range(len(work))]
    work["ID_PERSONA_MES"] = work[col_rut].astype(str).fillna("") + "||" + work[col_mes].astype(str).fillna("")
    work["MOVIMIENTO"] = "NETO_ORIGEN"
    work["MONTO_FINANCIADO"] = work["HABER_NETO"]
    work["ID_RELACION"] = work["ID_ORIGEN_FILA"]

    # Auditoría neto negativo
    neto_neg_mask = work["HABER_NETO"] < 0
    neto_negative_sample = work.loc[
        neto_neg_mask,
        [col_rut, col_mes, col_total_haber, "DESCUENTOS_BLOQUE", "HABER_NETO", col_cc, col_prog],
    ].head(50)
    neto_negative_count = int(neto_neg_mask.sum())

    # 7) Crear filas DAP (para descuentos > 0, incluso neto=0)
    has_discount = work["DESCUENTOS_REDISTRIBUIBLES"] > 0
    dap = work.loc[has_discount].copy()

    # Orden para insertar debajo
    work["_ORDER"] = np.arange(len(work)) * 2
    dap["_ORDER"] = work.loc[has_discount, "_ORDER"].values + 1

    # Config DAP
    dap[col_cc] = cc_dap
    dap[col_prog] = prog_dap
    dap[col_proceso] = "Redistribución Descuento a DAP"

    # blindaje anti-magnificación
    dap["TOTAL_HABER_BRUTO"] = 0.0
    dap[col_total_haber] = 0.0

    # en DAP: “neto” lo usamos como monto imputado (visual); lo oficial es MONTO_FINANCIADO
    dap["HABER_NETO"] = dap["DESCUENTOS_REDISTRIBUIBLES"]
    dap["MONTO_FINANCIADO"] = dap["DESCUENTOS_REDISTRIBUIBLES"]
    dap["MOVIMIENTO"] = "DESCUENTO_DAP"

    # esta fila no “aplica” descuentos; los recibe
    dap["DESCUENTOS_BLOQUE"] = 0.0
    dap["DESCUENTOS_REDISTRIBUIBLES"] = 0.0
    dap["DESCUENTOS_NO_REDISTRIB"] = 0.0
    dap["ID_RELACION"] = dap["ID_ORIGEN_FILA"]

    out = pd.concat([work, dap], ignore_index=False)
    out = out.sort_values("_ORDER", kind="stable").drop(columns=["_ORDER"])

    # Evita warning de fragmentación y mejora performance en operaciones siguientes
    out = out.copy()

    # 8) Resumen por (CC, Programa)
    resumen = (
        out.groupby([col_cc, col_prog], dropna=False)
        .agg(
            MONTO_FINANCIADO_TOTAL=("MONTO_FINANCIADO", "sum"),
            TOTAL_HABER_BRUTO_TOTAL=("TOTAL_HABER_BRUTO", "sum"),
            HABER_NETO_TOTAL=("HABER_NETO", "sum"),
            N_FILAS=("MOVIMIENTO", "count"),
            N_PERSONAS=(col_rut, pd.Series.nunique),
        )
        .reset_index()
        .sort_values("MONTO_FINANCIADO_TOTAL", ascending=False)
    )

    # 9) Estado de resultados por mes
    is_dap = (out[col_cc].astype(str) == cc_dap) & (_contains_ci(out[col_prog], "D.A.P"))
    grupo = np.where(is_dap, "DAP", "PRAPS")

    er = (
        out.assign(_GRUPO=grupo)
        .groupby([col_mes, "_GRUPO", "MOVIMIENTO"], dropna=False)
        .agg(MONTO_FINANCIADO=("MONTO_FINANCIADO", "sum"))
        .reset_index()
    )

    er_pivot = er.pivot_table(
        index=[col_mes],
        columns=["_GRUPO", "MOVIMIENTO"],
        values="MONTO_FINANCIADO",
        aggfunc="sum",
        fill_value=0.0,
    )

    er_pivot[("TOTAL", "FINANCIADO")] = er_pivot.sum(axis=1)
    if ("DAP", "DESCUENTO_DAP") in er_pivot.columns:
        denom = er_pivot[("TOTAL", "FINANCIADO")].replace({0: np.nan})
        er_pivot[("RATIO", "DAP_DESCUENTO_SOBRE_TOTAL")] = er_pivot[("DAP", "DESCUENTO_DAP")] / denom
    else:
        er_pivot[("RATIO", "DAP_DESCUENTO_SOBRE_TOTAL")] = np.nan

    er_pivot.columns = [f"{a}__{b}" for a, b in er_pivot.columns.to_list()]
    estado_resultados = er_pivot.reset_index()

    # Auditoría reintegros no pareados
    rein_unpaired_sample = rein_unpaired.copy()
    if not rein_unpaired_sample.empty:
        cols_show = [c for c in [col_proceso, col_rut, col_mes, col_total_haber, col_cc, col_prog] if c in rein_unpaired_sample.columns]
        rein_unpaired_sample = rein_unpaired_sample[cols_show].head(50)

    audit = AuditReport(
        missing_discount_columns=missing_cols,
        duplicate_discount_keys_in_block=dup_block_keys,
        reintegro_pairs_deleted=rein_pairs_deleted,
        reintegro_unpaired_rows=int(len(rein_unpaired)),
        reintegro_unpaired_sample=rein_unpaired_sample,
        neto_negative_rows=neto_negative_count,
        neto_negative_sample=neto_negative_sample,
    )

    return out, resumen, estado_resultados, audit








# Columnas nuevas del motor — deben tener nombre en fila 3, igual que las originales
_NEW_COLS = {
    "TOTAL_HABER_BRUTO", "DESCUENTOS_BLOQUE", "DESCUENTOS_REDISTRIBUIBLES",
    "DESCUENTOS_NO_REDISTRIB", "HABER_NETO", "ID_ORIGEN_FILA",
    "ID_PERSONA_MES", "MOVIMIENTO", "MONTO_FINANCIADO", "ID_RELACION",
}


def _fix_consolidado_headers(ws) -> None:
    """
    to_excel escribe: fila 1 = encabezados, fila 2+ = datos.
    El formato requerido es:  fila 1 = encabezados compuestos parte 1,
                              fila 2 = parte 2, fila 3 = nombres reales, fila 4+ = datos.
    Solución:
      1) Insertar 2 filas vacías al inicio → encabezados quedan en fila 3, datos en fila 4+
      2) Para columnas originales (NAN|NAN|X, INDICADOR|N|X, etc.):
         reconstruir fila 1 (parte 1) y fila 2 (parte 2) desde el nombre compuesto
         y dejar fila 3 con el nombre real (parte 3)
      3) Para columnas nuevas (MOVIMIENTO, HABER_NETO, etc.):
         fila 1 y 2 quedan vacías, fila 3 = nombre → ya correcto tras la inserción
    """
    # Paso 1: insertar 2 filas vacías al inicio
    ws.insert_rows(1, amount=2)
    # Ahora: fila 1=vacía, fila 2=vacía, fila 3=encabezados originales, fila 4+=datos

    # Paso 2: reconstruir fila 1 y fila 2 para columnas con encabezado compuesto
    for cell in ws[3]:  # fila 3 tiene todos los nombres tal como los escribió to_excel
        col_name = cell.value
        if col_name is None or col_name in _NEW_COLS:
            continue  # columnas nuevas: filas 1 y 2 vacías — correcto
        # Columna original compuesta: "PARTE1 | PARTE2 | PARTE3"
        parts = [p.strip() for p in str(col_name).split("|")]
        if len(parts) == 3:
            p1 = None if parts[0].upper() in ("NAN", "NONE", "") else parts[0]
            p2 = None if parts[1].upper() in ("NAN", "NONE", "") else parts[1]
            p3 = parts[2] if parts[2].upper() not in ("NAN", "NONE", "") else parts[0]
        else:
            p1 = None; p2 = None; p3 = col_name
        col = cell.column
        ws.cell(row=1, column=col).value = p1
        ws.cell(row=2, column=col).value = p2
        ws.cell(row=3, column=col).value = p3  # nombre real limpio

# =========================
# Escritura Excel + Charts
# =========================
def write_output_excel(
    output_path: str,
    consolidado_out: pd.DataFrame,
    resumen: pd.DataFrame,
    estado_resultados: pd.DataFrame,
    audit: AuditReport,
) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        consolidado_out.to_excel(writer, index=False, sheet_name="CONSOLIDADO_REDISTRIBUIDO")
        _fix_consolidado_headers(writer.book["CONSOLIDADO_REDISTRIBUIDO"])
        resumen.to_excel(writer, index=False, sheet_name="RESUMEN_CC_PROGRAMA")
        estado_resultados.to_excel(writer, index=False, sheet_name="ESTADO_RESULTADOS")

        # AUDITORIA
        audit_rows = [
            ["reintegro_pairs_deleted", audit.reintegro_pairs_deleted],
            ["reintegro_unpaired_rows", audit.reintegro_unpaired_rows],
            ["neto_negative_rows", audit.neto_negative_rows],
            ["missing_discount_columns_count", len(audit.missing_discount_columns)],
            ["duplicate_discount_keys_in_block_count", len(audit.duplicate_discount_keys_in_block)],
        ]
        audit_df = pd.DataFrame(audit_rows, columns=["METRICA", "VALOR"])
        audit_df.to_excel(writer, index=False, sheet_name="AUDITORIA", startrow=0)

        start = len(audit_df) + 3
        pd.DataFrame({"MISSING_DISCOUNT_KEYS": audit.missing_discount_columns}).to_excel(
            writer, index=False, sheet_name="AUDITORIA", startrow=start
        )

        start2 = start + max(3, len(audit.missing_discount_columns) + 3)
        pd.DataFrame({"DUPLICATE_BLOCK_KEYS": audit.duplicate_discount_keys_in_block}).to_excel(
            writer, index=False, sheet_name="AUDITORIA", startrow=start2
        )

        start3 = start2 + max(3, len(audit.duplicate_discount_keys_in_block) + 3)
        if audit.reintegro_unpaired_sample is not None and not audit.reintegro_unpaired_sample.empty:
            audit.reintegro_unpaired_sample.to_excel(
                writer, index=False, sheet_name="AUDITORIA", startrow=start3
            )
            start3 += len(audit.reintegro_unpaired_sample) + 3

        if audit.neto_negative_sample is not None and not audit.neto_negative_sample.empty:
            audit.neto_negative_sample.to_excel(
                writer, index=False, sheet_name="AUDITORIA", startrow=start3
            )

    # Gráficos
    wb = load_workbook(output_path)
    if "ESTADO_RESULTADOS" in wb.sheetnames:
        ws = wb["ESTADO_RESULTADOS"]
        headers = [cell.value for cell in ws[1]]

        def col_index(name: str) -> Optional[int]:
            try:
                return headers.index(name) + 1
            except ValueError:
                return None

        mes_col = col_index("MES PAGO") or col_index("MES") or 1
        praps_neto = col_index("PRAPS__NETO_ORIGEN")
        dap_desc = col_index("DAP__DESCUENTO_DAP")
        total_fin = col_index("TOTAL__FINANCIADO")

        max_row = ws.max_row
        if max_row >= 3 and (praps_neto or dap_desc):
            chart = BarChart()
            chart.type = "col"
            chart.grouping = "stacked"
            chart.title = "Redistribución: PRAPS (Neto) vs DAP (Descuentos)"
            chart.y_axis.title = "Monto (pesos)"
            chart.x_axis.title = "Mes pago"
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = False

            series_cols = [c for c in [praps_neto, dap_desc] if c is not None]
            data_ref = Reference(ws, min_col=min(series_cols), max_col=max(series_cols), min_row=1, max_row=max_row)
            cats_ref = Reference(ws, min_col=mes_col, min_row=2, max_row=max_row)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            ws.add_chart(chart, "H2")

        if max_row >= 3 and total_fin is not None:
            chart2 = BarChart()
            chart2.type = "col"
            chart2.title = "Total financiado por mes (MONTO_FINANCIADO)"
            chart2.y_axis.title = "Monto (pesos)"
            chart2.x_axis.title = "Mes pago"
            data_ref2 = Reference(ws, min_col=total_fin, max_col=total_fin, min_row=1, max_row=max_row)
            cats_ref2 = Reference(ws, min_col=mes_col, min_row=2, max_row=max_row)
            chart2.add_data(data_ref2, titles_from_data=True)
            chart2.set_categories(cats_ref2)
            ws.add_chart(chart2, "H20")

    wb.save(output_path)


# =========================
# UI Streamlit
# =========================
def run_streamlit_app():
    import traceback
    import streamlit as st

    st.set_page_config(page_title="Redistribución PRAPS vs DAP", layout="wide")
    st.title("Redistribución PRAPS vs D.A.P (ADMINISTRACION)")
    st.caption("Sube el Excel, selecciona las hojas y presiona **GENERAR REDISTRIBUCIÓN**.")

    uploaded = st.file_uploader("Sube el archivo Excel (.xlsx)", type=["xlsx"])

    if not uploaded:
        st.info("Esperando archivo...")
        st.stop()

    # Leer hojas (liviano)
    try:
        xls = pd.ExcelFile(uploaded, engine="openpyxl")
        sheets = xls.sheet_names
    except Exception:
        st.error("No pude leer el Excel. ¿Está corrupto o protegido con contraseña?")
        st.code(traceback.format_exc())
        st.stop()

    col1, col2 = st.columns([2, 1])
    with col1:
        sheet_consolidado = st.selectbox(
            "Hoja CONSOLIDADO",
            options=sheets,
            index=sheets.index("CONSOLIDADO") if "CONSOLIDADO" in sheets else 0,
        )
        sheet_bloque = st.selectbox(
            "Hoja BLOQUE DESCUENTO",
            options=sheets,
            index=sheets.index("BLOQUE DESCUENTO") if "BLOQUE DESCUENTO" in sheets else 0,
        )
    with col2:
        header_rows = st.number_input(
            "Filas encabezado compuesto (CONSOLIDADO)",
            min_value=1,
            max_value=6,
            value=3,
            step=1,
        )

    st.divider()

    if st.button("GENERAR REDISTRIBUCIÓN", type="primary", use_container_width=True):
        try:
            with st.spinner("Procesando... (esto puede tardar si el Excel es muy grande)"):
                uploaded.seek(0)

                consolidado_out, resumen, estado, audit = generate_redistribution(
                    input_path_or_buffer=uploaded,
                    sheet_consolidado=sheet_consolidado,
                    sheet_bloque=sheet_bloque,
                    header_rows=int(header_rows),
                )

                # Guardar en archivo temporal y devolver bytes
                import tempfile

                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                    tmp_path = tmp.name

                write_output_excel(tmp_path, consolidado_out, resumen, estado, audit)

                with open(tmp_path, "rb") as f:
                    out_bytes = f.read()

                try:
                    os.remove(tmp_path)
                except Exception:
                    pass

            st.success("Listo. Descarga el archivo resultante:")
            st.download_button(
                "Descargar SALIDA_REDISTRIBUCION.xlsx",
                data=out_bytes,
                file_name="SALIDA_REDISTRIBUCION.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

            st.subheader("Auditoría rápida")
            st.write(
                {
                    "Pares reintegro borrados": audit.reintegro_pairs_deleted,
                    "Reintegros no pareados": audit.reintegro_unpaired_rows,
                    "Neto negativo (casos)": audit.neto_negative_rows,
                    "Conceptos bloque NO encontrados": len(audit.missing_discount_columns),
                    "Llaves duplicadas en bloque": len(audit.duplicate_discount_keys_in_block),
                }
            )

            st.subheader("Vista previa: Resumen por (Centro de costo, Programa)")
            st.dataframe(resumen.head(50), use_container_width=True)

            if audit.missing_discount_columns:
                st.warning("Hay conceptos del BLOQUE DESCUENTO que NO se encontraron en el CONSOLIDADO (llave compuesta).")
                st.dataframe(pd.DataFrame({"MISSING_DISCOUNT_KEYS": audit.missing_discount_columns}).head(100))

        except Exception:
            st.error("La ejecución falló. Copia este traceback y pégamelo:")
            st.code(traceback.format_exc())
            st.stop()


# =========================
# CLI opcional (NO rompe Streamlit)
# =========================
def main_cli():
    p = argparse.ArgumentParser()
    p.add_argument("--input", required=False, help="Ruta al archivo Excel .xlsx")
    p.add_argument("--output", required=False, help="Ruta salida .xlsx (default: SALIDA_REDISTRIBUCION.xlsx)")
    p.add_argument("--sheet_consolidado", default="CONSOLIDADO")
    p.add_argument("--sheet_bloque", default="BLOQUE DESCUENTO")
    p.add_argument("--header_rows", type=int, default=3)
    args, _unknown = p.parse_known_args()

    if not args.input:
        print("Error: Debes indicar --input (modo CLI). Para UI usa: streamlit run app_redistribucion.py", file=sys.stderr)
        return

    output = args.output or "SALIDA_REDISTRIBUCION.xlsx"
    consolidado_out, resumen, estado, audit = generate_redistribution(
        input_path_or_buffer=args.input,
        sheet_consolidado=args.sheet_consolidado,
        sheet_bloque=args.sheet_bloque,
        header_rows=args.header_rows,
    )
    write_output_excel(output, consolidado_out, resumen, estado, audit)
    print(f"OK -> {output}")
    print(
        f"Auditoría: pares reintegro borrados={audit.reintegro_pairs_deleted}, "
        f"reintegros no pareados={audit.reintegro_unpaired_rows}, "
        f"neto negativo={audit.neto_negative_rows}, "
        f"missing bloque={len(audit.missing_discount_columns)}"
    )


# =========================
# Entry point
# =========================
if __name__ == "__main__":
    # Si lo ejecutas con streamlit run, Streamlit ejecuta el script y la UI se renderiza desde arriba.
    # Si lo ejecutas con python, se usa CLI.
    if running_under_streamlit():
        run_streamlit_app()
    else:
        main_cli()
