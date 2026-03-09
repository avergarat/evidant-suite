# pip install openpyxl

import threading
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter


# ----------------------------
# Utilidades encabezados
# ----------------------------

def norm(x):
    if x is None:
        return ""
    s = str(x).strip()
    # normalización suave (si quieres más agresivo: upper(), reemplazos, etc.)
    return s

def find_header_rows(ws, scan_rows=80):
    """
    Encuentra la fila del header3 buscando la fila que contiene 'PROCESO' (normalmente columna A).
    Retorna (header1_row, header2_row, header3_row).
    """
    # Escanea primeras N filas
    for r in range(1, scan_rows + 1):
        v1 = norm(ws.cell(row=r, column=1).value)
        # Heurística principal: columna A = 'PROCESO'
        if v1.upper() == "PROCESO":
            if r < 3:
                raise ValueError(f"Header3 encontrado en fila {r}, pero no hay 2 filas arriba para header1/header2.")
            return (r - 2, r - 1, r)

    # Heurística alternativa: buscar fila donde existan PROCESO y FOLIO en primeras columnas
    for r in range(1, scan_rows + 1):
        row_vals = [norm(ws.cell(row=r, column=c).value).upper() for c in range(1, 20)]
        if "PROCESO" in row_vals and "FOLIO" in row_vals:
            if r < 3:
                raise ValueError(f"Header3 encontrado en fila {r}, pero no hay 2 filas arriba para header1/header2.")
            return (r - 2, r - 1, r)

    raise ValueError("No pude detectar el header3 (fila con 'PROCESO') en las primeras filas. Ajusta scan_rows o revisa la hoja.")


def get_max_col_smart(ws, header3_row, scan_extra=50):
    """
    Estima la última columna a considerar usando la fila header3 (nombres de columna)
    más un margen por si hay columnas con header vacío pero datos.
    """
    # ws.max_column puede estar inflado por estilos; en read_only suele ser más razonable,
    # igual lo acotamos a algo basado en lo visible cerca del header.
    max_col = ws.max_column

    # Busca última columna no vacía en header3
    last_nonempty = 0
    for c in range(1, max_col + 1):
        if norm(ws.cell(row=header3_row, column=c).value) != "":
            last_nonempty = c

    # Margen: si hay datos a la derecha con header vacío, esto los puede capturar
    return max(last_nonempty + scan_extra, last_nonempty, 1)


def build_sheet_header_keys(ws, h1, h2, h3):
    """
    Construye lista de claves por columna (triplete).
    Si el triplete completo está vacío, igual crea una clave única por posición para no perder la columna.
    """
    max_col = get_max_col_smart(ws, h3)
    keys = []
    for c in range(1, max_col + 1):
        a = norm(ws.cell(row=h1, column=c).value)
        b = norm(ws.cell(row=h2, column=c).value)
        d = norm(ws.cell(row=h3, column=c).value)
        if a == "" and b == "" and d == "":
            # columna sin encabezados: no la mezclamos con otras; se preserva por posición
            keys.append((f"__BLANK__{get_column_letter(c)}", "", ""))
        else:
            keys.append((a, b, d))
    return keys


# ----------------------------
# Consolidación (streaming)
# ----------------------------

def consolidate_by_header_triplet(input_path, output_path, sheet_names, add_source_col=True):
    """
    Une columnas por (header1, header2, header3). Filas se apilan.
    Escribe salida con 3 filas de encabezado + (opcional) ORIGEN_HOJA.
    """
    wb = load_workbook(input_path, read_only=True, data_only=False, keep_links=False)

    # 1) Detectar headers y construir orden global de columnas (unión preservando orden)
    global_cols = []          # lista de tripletes
    global_index = {}         # triplete -> idx
    per_sheet = {}            # info por hoja: header rows, keys, colmap, start_data_row, max_col

    for sname in sheet_names:
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        h1, h2, h3 = find_header_rows(ws)

        keys = build_sheet_header_keys(ws, h1, h2, h3)
        max_col = len(keys)

        # Crear mapping col origen -> col global
        colmap = {}
        for c, k in enumerate(keys, start=1):
            if k not in global_index:
                global_index[k] = len(global_cols)
                global_cols.append(k)
            colmap[c] = global_index[k]

        per_sheet[sname] = {
            "h1": h1, "h2": h2, "h3": h3,
            "start_data_row": h3 + 1,
            "max_col": max_col,
            "colmap": colmap
        }

    if not global_cols:
        raise ValueError("No se pudo construir ninguna columna global. Revisa nombres/selección de hojas.")

    # 2) Crear salida write_only (baja RAM)
    wb_out = Workbook(write_only=True)
    ws_out = wb_out.create_sheet("CONSOLIDADO")

    # Columnas finales: opcional ORIGEN_HOJA al inicio
    offset = 1 if add_source_col else 0

    # 3) Escribir 3 filas de encabezado
    row_h1 = (["ORIGEN_HOJA"] if add_source_col else []) + [k[0] for k in global_cols]
    row_h2 = ([""] if add_source_col else []) + [k[1] for k in global_cols]
    row_h3 = ([""] if add_source_col else []) + [k[2] for k in global_cols]

    ws_out.append(row_h1)
    ws_out.append(row_h2)
    ws_out.append(row_h3)

    # 4) Escribir filas de datos apiladas
    for sname in sheet_names:
        if sname not in per_sheet:
            continue
        ws = wb[sname]
        info = per_sheet[sname]
        start_r = info["start_data_row"]
        max_col = info["max_col"]
        colmap = info["colmap"]

        # Streaming: iter_rows desde start_r
        for row in ws.iter_rows(min_row=start_r, max_col=max_col, values_only=True):
            # Si la fila es completamente vacía, puedes saltarla para no inflar salida
            if row is None:
                continue
            if all((v is None or v == "") for v in row):
                continue

            out = [None] * (len(global_cols) + offset)
            if add_source_col:
                out[0] = sname

            # mapear valores
            for c, v in enumerate(row, start=1):
                gi = colmap.get(c)
                if gi is None:
                    continue
                out[gi + offset] = v

            ws_out.append(out)

    wb_out.save(output_path)


# ----------------------------
# FRONT Tkinter
# ----------------------------

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Consolidar por triplete de encabezados (3 filas)")
        self.geometry("820x520")

        self.input_path = tk.StringVar()
        self.output_path = tk.StringVar()
        self.add_source = tk.BooleanVar(value=True)
        self.status = tk.StringVar(value="Listo.")

        frm = tk.Frame(self, padx=12, pady=12)
        frm.pack(fill="both", expand=True)

        tk.Label(frm, text="Archivo Excel de entrada:").grid(row=0, column=0, sticky="w")
        tk.Entry(frm, textvariable=self.input_path, width=80).grid(row=1, column=0, sticky="we")
        tk.Button(frm, text="Abrir...", command=self.pick_input).grid(row=1, column=1, padx=8)

        tk.Label(frm, text="Hojas a consolidar (en este orden):").grid(row=2, column=0, sticky="w", pady=(12, 0))
        self.listbox = tk.Listbox(frm, selectmode="extended", height=12)
        self.listbox.grid(row=3, column=0, sticky="nsew")

        btns = tk.Frame(frm)
        btns.grid(row=3, column=1, sticky="n", padx=8)
        tk.Button(btns, text="Subir", width=12, command=self.move_up).pack(pady=2)
        tk.Button(btns, text="Bajar", width=12, command=self.move_down).pack(pady=2)
        tk.Button(btns, text="Seleccionar\ntodo", width=12, command=self.select_all).pack(pady=8)

        tk.Checkbutton(frm, text="Agregar columna ORIGEN_HOJA (recomendado)", variable=self.add_source)\
            .grid(row=4, column=0, sticky="w", pady=(10, 0))

        tk.Label(frm, text="Archivo de salida:").grid(row=5, column=0, sticky="w", pady=(12, 0))
        tk.Entry(frm, textvariable=self.output_path, width=80).grid(row=6, column=0, sticky="we")
        tk.Button(frm, text="Guardar como...", command=self.pick_output).grid(row=6, column=1, padx=8)

        tk.Label(frm, textvariable=self.status).grid(row=7, column=0, sticky="w", pady=(10, 0))

        self.btn_run = tk.Button(frm, text="CONSOLIDAR", command=self.run_threaded, height=2)
        self.btn_run.grid(row=8, column=0, sticky="we", pady=12)

        frm.grid_columnconfigure(0, weight=1)
        frm.grid_rowconfigure(3, weight=1)

    def pick_input(self):
        path = filedialog.askopenfilename(
            title="Selecciona Excel",
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")]
        )
        if not path:
            return
        self.input_path.set(path)

        try:
            wb = load_workbook(path, read_only=True, data_only=False, keep_links=False)
        except Exception as e:
            messagebox.showerror("Error", f"No pude abrir el archivo:\n{e}")
            return

        self.listbox.delete(0, tk.END)
        for n in wb.sheetnames:
            self.listbox.insert(tk.END, n)

        for i in range(min(3, self.listbox.size())):
            self.listbox.selection_set(i)

        if not self.output_path.get():
            self.output_path.set(path.replace(".xlsx", "_CONSOLIDADO_UNION.xlsx"))

    def pick_output(self):
        path = filedialog.asksaveasfilename(
            title="Guardar consolidado como...",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )
        if path:
            self.output_path.set(path)

    def select_all(self):
        self.listbox.selection_set(0, tk.END)

    def move_up(self):
        sel = list(self.listbox.curselection())
        if not sel:
            return
        for idx in sel:
            if idx == 0:
                continue
            text = self.listbox.get(idx)
            self.listbox.delete(idx)
            self.listbox.insert(idx - 1, text)
            self.listbox.selection_set(idx - 1)

    def move_down(self):
        sel = list(self.listbox.curselection())
        if not sel:
            return
        for idx in reversed(sel):
            if idx == self.listbox.size() - 1:
                continue
            text = self.listbox.get(idx)
            self.listbox.delete(idx)
            self.listbox.insert(idx + 1, text)
            self.listbox.selection_set(idx + 1)

    def run_threaded(self):
        in_path = self.input_path.get().strip()
        out_path = self.output_path.get().strip()
        sel = list(self.listbox.curselection())

        if not in_path:
            messagebox.showwarning("Falta archivo", "Selecciona el archivo de entrada.")
            return
        if not out_path:
            messagebox.showwarning("Falta salida", "Selecciona el archivo de salida.")
            return
        if not sel:
            messagebox.showwarning("Faltan hojas", "Selecciona al menos una hoja.")
            return

        sheet_names = [self.listbox.get(i) for i in sel]

        self.btn_run.config(state="disabled")
        self.status.set("Consolidando (unión por triplete de encabezados)...")

        def worker():
            try:
                consolidate_by_header_triplet(
                    input_path=in_path,
                    output_path=out_path,
                    sheet_names=sheet_names,
                    add_source_col=self.add_source.get()
                )
                self.after(0, lambda: messagebox.showinfo("Listo", f"Consolidado generado:\n{out_path}"))
                self.after(0, lambda: self.status.set("Listo."))
            except Exception as e:
                self.after(0, lambda: messagebox.showerror("Error", f"Falló:\n{e}"))
                self.after(0, lambda: self.status.set("Error."))
            finally:
                self.after(0, lambda: self.btn_run.config(state="normal"))

        threading.Thread(target=worker, daemon=True).start()


if __name__ == "__main__":
    App().mainloop()
