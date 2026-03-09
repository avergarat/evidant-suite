import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import threading
import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
import unicodedata

# --- Contratos cortos: reglas selectivas por llave (v18_5) ---
TARGET_SUPLENCIAS_SOLO_CORTOS: set[str] = {
    "111111019664",
    "121111019664",
    "211111018834",
    "221111018834",
}

LLAVES_SIN_CORTOS: set[str] = {
    "112115076","115215076","212118834","112119664","212218834","2117718834","212618834",
    "122119664","222118834","222218834","2217718834","222618834","223018834","216118834","226118834","2131618834","2231618834",
    "12518834","214718834","119719664","112519664","122519664","222518834","123419664","224718834","129619664","224818834","129719664",

    
}


from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string


# =========================
# Utils
# =========================
def parse_col_range(col_range: str) -> tuple[int, int]:
    s = col_range.strip().upper()
    m = re.match(r"^([A-Z]{1,3})\s*[:\-\s]\s*([A-Z]{1,3})$", s)
    if not m:
        raise ValueError("Rango inválido. Ejemplo: CK:EP")
    c1, c2 = m.group(1), m.group(2)
    i1 = column_index_from_string(c1)
    i2 = column_index_from_string(c2)
    return (i1, i2) if i1 <= i2 else (i2, i1)


def get_sheetnames_fast_xlsx(xlsx_path: str) -> list[str]:
    with zipfile.ZipFile(xlsx_path, "r") as z:
        wb_xml = z.read("xl/workbook.xml")
    root = ET.fromstring(wb_xml)

    def local(tag: str) -> str:
        return tag.split("}", 1)[1] if "}" in tag else tag

    names = []
    for el in root.iter():
        if local(el.tag) == "sheet":
            nm = el.attrib.get("name")
            if nm:
                names.append(nm)
    return names


def norm_text(x) -> str:
    if x is None:
        return ""
    s = str(x)
    s = s.replace("\u00A0", " ").replace("\t", " ").replace("\n", " ").strip()
    s = re.sub(r"\s+", " ", s)
    return s


def norm_header(x) -> str:
    s = norm_text(x).upper()
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"[^A-Z0-9 ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def digits_only(x) -> str:
    s = norm_text(x)
    if not s:
        return ""
    return re.sub(r"\D+", "", s)


def to_number(x):
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if s == "":
        return None
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s and "." not in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except:
        return None


def getv(row_vals, c: int):
    if row_vals is None:
        return None
    j = c - 1
    return row_vals[j] if 0 <= j < len(row_vals) else None


def find_header_row_and_map(ws, required_headers: list[str], search_rows: int = 80):
    req = [norm_header(h) for h in required_headers]
    max_r = min(ws.max_row or 1, search_rows)
    for r in range(1, max_r + 1):
        row = next(ws.iter_rows(min_row=r, max_row=r, values_only=True), None) or ()
        m = {}
        for idx, v in enumerate(row, start=1):
            k = norm_header(v)
            if k and k not in m:
                m[k] = idx
        if all(h in m for h in req):
            return r, m
    return None, None


def find_col_by_contains(header_map: dict, needle: str) -> int | None:
    n = norm_header(needle)
    for k, idx in header_map.items():
        if n in k:
            return idx
    return None


def find_col_exact(header_map: dict, exact_name: str) -> int | None:
    """Match exacto de encabezado normalizado. Evita confundir 'CODIGO' con 'CODIGO SIGFE'."""
    target = norm_header(exact_name)
    for k, idx in header_map.items():
        if k == target:
            return idx
    return None


def find_total_col(header_map: dict, ley: str | None = None) -> int | None:
    # En modo multi-ley no buscamos "TOTAL LEY 18834" etc, buscamos "TOTAL LEY" genérico
    if ley:
        ley_digits = re.sub(r"\D+", "", str(ley))
        if ley_digits:
            for k, idx in header_map.items():
                if k.startswith("TOTAL LEY") and ley_digits in k:
                    return idx
    for k, idx in header_map.items():
        if k.startswith("TOTAL LEY"):
            return idx
    for k, idx in header_map.items():
        if "TOTAL" in k and "LEY" in k:
            return idx
    return None


def normalize_month(v) -> int | None:
    n = to_number(v)
    if n is None:
        return None
    mi = int(n)
    if 1 <= mi <= 12:
        return mi
    return None


def norm_cc(v) -> str:
    s = norm_text(v)
    if not s:
        return ""
    return re.sub(r"\s+", " ", s).strip()


def norm_proc(v) -> str:
    return norm_cc(v)


def split_cj_calidad(v):
    """Separa '21_TITULARES' en ('21','TITULARES'). Si no calza: ('', texto_limpio)."""
    s = norm_text(v)
    if not s:
        return "", ""
    m = re.match(r"^\s*(\d+)\s*_(.+?)\s*$", s)
    if m:
        return m.group(1), m.group(2).strip()
    return "", s


# =========================
# Cache simple en memoria para acelerar recargas (UI)
# =========================
_LIST_CACHE: dict[tuple[str, str, str], list[str]] = {}


def _get_unique_values_from_base(
    src_path: str,
    base_sheet_name: str,
    header_name: str,
    search_rows_header: int = 200,
    max_scan_rows: int = 200000,
    normalizer=None,
) -> list[str]:
    """Lee valores únicos de una sola columna en BASE de forma rápida (read_only + solo 1 columna)."""
    key = (str(Path(src_path).resolve()), base_sheet_name, norm_header(header_name))
    if key in _LIST_CACHE:
        return list(_LIST_CACHE[key])

    wb = load_workbook(src_path, read_only=True, data_only=True)
    try:
        if base_sheet_name not in wb.sheetnames:
            raise ValueError(f'No existe hoja BASE "{base_sheet_name}".')
        ws = wb[base_sheet_name]

        hdr_row, hdr_map = find_header_row_and_map(ws, [header_name], search_rows=search_rows_header)
        if hdr_row is None:
            raise ValueError(f'No pude detectar "{header_name}" en BASE (primeras {search_rows_header} filas).')

        col = find_col_by_contains(hdr_map, header_name)
        if col is None:
            raise ValueError(f'Detecté encabezados, pero no pude ubicar columna "{header_name}".')

        data_start = hdr_row + 1
        max_r = min(ws.max_row or 1, max_scan_rows)

        seen = set()
        out = []
        for (v,) in ws.iter_rows(min_row=data_start, max_row=max_r, min_col=col, max_col=col, values_only=True):
            s = normalizer(v) if normalizer else norm_text(v)
            if s and s not in seen:
                seen.add(s)
                out.append(s)

        _LIST_CACHE[key] = list(out)
        return out
    finally:
        wb.close()


def get_centros_costo_from_base_xlsx(src_path: str, base_sheet_name: str) -> list[str]:
    return _get_unique_values_from_base(src_path, base_sheet_name, "CENTRO DE COSTO", normalizer=norm_cc)


def get_procesos_from_base_xlsx(src_path: str, base_sheet_name: str) -> list[str]:
    return _get_unique_values_from_base(src_path, base_sheet_name, "PROCESO", normalizer=norm_proc)


# =========================
# CORE
# =========================
def process_steps_1_to_4(
    src_path: str,
    base_sheet_name: str,
    consol_sheet_name: str,
    col_range_montos: str,
    law: str,                         # fallback interno (NO cambia reglas si LEY AFECTO existe)
    laws_selected: list[str] | None = None,  # selección real para filtrar (multi-ley)
    out_path: str | None = None,
    months_selected: list[int] | None = None,
    centros_selected: list[str] | None = None,
    procesos_selected: list[str] | None = None,
    generate_pf: bool = False,
    pf_model_sheet_name: str | None = None,
) -> str:
    src = Path(src_path)
    if out_path is None:
        out_path = str(src.with_name(f"{src.stem}_SALIDA{src.suffix}"))

    # meses
    if months_selected is None or len(months_selected) == 0:
        selected_months = list(range(1, 13))
    else:
        selected_months = sorted({int(m) for m in months_selected if 1 <= int(m) <= 12})

    # centros
    selected_centros = None
    selected_centros_ordered: list[str] | None = None
    if centros_selected is not None and len(centros_selected) > 0:
        cleaned = [norm_cc(x) for x in centros_selected if norm_cc(x)]
        selected_centros_ordered = cleaned
        selected_centros = set(cleaned)

    # procesos
    selected_procesos = None
    if procesos_selected is not None and len(procesos_selected) > 0:
        selected_procesos = {norm_proc(x) for x in procesos_selected if norm_proc(x)}

    # filtro de leyes (multi)
    laws_filter = None
    if laws_selected:
        laws_filter = {digits_only(l) for l in laws_selected if digits_only(l)}
        if not laws_filter:
            laws_filter = None

    law_fallback = digits_only(law) or norm_text(law)

    wb_src = load_workbook(src_path, read_only=True, data_only=True)
    try:
        if base_sheet_name not in wb_src.sheetnames:
            raise ValueError(f'No existe hoja BASE "{base_sheet_name}".')
        if consol_sheet_name not in wb_src.sheetnames:
            raise ValueError(f'No existe hoja CONSOLIDACIÓN "{consol_sheet_name}".')

        if generate_pf:
            if not pf_model_sheet_name:
                raise ValueError('Debes seleccionar la hoja "PROGRAMA FINANCIERO (modelo)".')
            if pf_model_sheet_name not in wb_src.sheetnames:
                raise ValueError(f'No existe hoja MODELO "{pf_model_sheet_name}".')

        ws_base = wb_src[base_sheet_name]
        ws_consol = wb_src[consol_sheet_name]
        ws_pf_model = wb_src[pf_model_sheet_name] if generate_pf else None

        max_col_base = ws_base.max_column or 1
        max_row_base = ws_base.max_row or 1

        # ===== rango dinámico por metadata fila 1 =====
        def _detect_dynamic_range_from_row1() -> tuple[int, int]:
            if max_row_base < 1:
                raise ValueError('La hoja BASE no tiene filas suficientes (fila 1).')
            row1_vals = next(ws_base.iter_rows(min_row=1, max_row=1, values_only=True), None) or ()
            first = None
            last = None
            for c in range(1, max_col_base + 1):
                if norm_header(getv(row1_vals, c)):
                    first = c
                    break
            for c in range(max_col_base, 0, -1):
                if norm_header(getv(row1_vals, c)):
                    last = c
                    break
            if first is None or last is None or first > last:
                raise ValueError('No pude detectar rango dinámico: la fila 1 no contiene metadata utilizable.')
            return first, last

        start_col, end_col = _detect_dynamic_range_from_row1()

        # ===== headers BASE =====
        base_header_row, base_map = find_header_row_and_map(
            ws_base, ["MES PAGO", "CENTRO DE COSTO", "PROCESO"], search_rows=220
        )
        if base_header_row is None:
            raise ValueError('No pude detectar encabezados en BASE: necesito "PROCESO", "MES PAGO", "CENTRO DE COSTO".')

        cj_col_idx = find_col_by_contains(base_map, "C/J")
        calidad_col_idx = find_col_by_contains(base_map, "CALIDAD JURIDICA")
        mes_col_idx = find_col_by_contains(base_map, "MES PAGO")
        cc_col_idx = find_col_by_contains(base_map, "CENTRO DE COSTO")
        pr_col_idx = find_col_by_contains(base_map, "PROCESO")
        rut_col_idx = find_col_by_contains(base_map, "RUT")  # soporta 'RUT-DV'
        ley_afecto_col_idx = find_col_by_contains(base_map, "LEY AFECTO")

        if cj_col_idx is None and calidad_col_idx is None:
            raise ValueError('No encontré "C/J" ni "CALIDAD JURIDICA" en BASE (necesito al menos una).')
        if mes_col_idx is None:
            raise ValueError('No pude ubicar "MES PAGO" en BASE.')
        if cc_col_idx is None:
            raise ValueError('No pude ubicar "CENTRO DE COSTO" en BASE.')
        if pr_col_idx is None:
            raise ValueError('No pude ubicar "PROCESO" en BASE.')
        if rut_col_idx is None:
            raise ValueError('No pude ubicar "RUT-DV" (o "RUT") en BASE. Es necesaria para Aporte Bienestar.')

        base_data_start = base_header_row + 1

        # ===== insertar C/J si falta =====
        cj_missing_in_source = (cj_col_idx is None)
        cj_insert_pos = None
        if cj_missing_in_source:
            cj_insert_pos = calidad_col_idx if calidad_col_idx is not None else (max_col_base + 1)

        # ===== Paso 0: detectar columnas INDICADOR (fila 1) y codes (fila 2) =====
        indicator_cols: set[int] = set()
        if max_row_base >= 1:
            row1 = next(ws_base.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if row1:
                for c in range(start_col, end_col + 1):
                    h1 = norm_header(getv(row1, c))
                    if h1 == "INDICADOR":
                        indicator_cols.add(c)

        # ===== Paso 0: detectar columnas INDICADOR (fila 1) y codes (fila 2) =====
        indicator_cols: set[int] = set()
        if max_row_base >= 1:
            row1 = next(ws_base.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if row1:
                for c in range(start_col, end_col + 1):
                    h1 = norm_header(getv(row1, c))
                    if h1 == "INDICADOR":
                        indicator_cols.add(c)

        # En vez de codes[col]=code, guardamos code -> [cols...]
        code_to_cols: dict[str, list[int]] = {}
        if max_row_base >= 2:
            row2 = next(ws_base.iter_rows(min_row=2, max_row=2, values_only=True), None)
            if row2:
                for c in range(start_col, end_col + 1):
                    v = getv(row2, c)
                    s = digits_only(v)
                    if not s:
                        continue
                    if c in indicator_cols and not s.startswith("111"):
                        s = f"111{s}"
                    code_to_cols.setdefault(s, []).append(c)

        # columnas especiales (tomamos la primera ocurrencia)
        contrato_corto_col = (code_to_cols.get("11110") or [None])[0]  # indicador 10
        indicador_52_col   = (code_to_cols.get("11152") or [None])[0]  # indicador 52


        wb_out = Workbook(write_only=True)
        ws_out_base = wb_out.create_sheet(title=base_sheet_name)
        ws_out_con = wb_out.create_sheet(title=consol_sheet_name)

        # acumulación por (llave, mes) filtrada por CC/PROCESO/LEY
        sums_by_llave_month: dict[str, dict[int, float]] = {}
        sums_by_llave_month_S: dict[str, dict[int, float]] = {}
        bienestar_ruts_by_llave_month: dict[str, dict[int, set[str]]] = {}
        sums_by_llave_month_cc: dict[str, dict[int, dict[str, float]]] = {}
        sums_by_llave_month_N: dict[str, dict[int, float]] = {}
        sums_by_llave_month_cc_all: dict[str, dict[int, dict[str, float]]] = {}
        sums_by_llave_month_cc_S: dict[str, dict[int, dict[str, float]]] = {}

        in_range = [False] * (max_col_base + 1)
        for c in range(start_col, end_col + 1):
            if 1 <= c <= max_col_base:
                in_range[c] = True

        detected_cc_order: list[str] = []
        detected_cc_set: set[str] = set()

        # ===== BASE stream + acumulación =====
        for r_idx, row_vals in enumerate(ws_base.iter_rows(values_only=True), start=1):
            out_row = []

            mes_pago = None
            cc_val = ""
            pr_val = ""
            cj_digits = ""
            calidad_txt = None
            contrato_corto_flag = ""  # 'N' o 'S'
            ind52_flag = ""           # 'SI' o 'NO'
            rut_dv = ""
            row_law_digits = ""

            passes_filters_row = False
            passes_law_row = True

            if r_idx >= base_data_start:
                mes_pago = normalize_month(getv(row_vals, mes_col_idx))
                cc_val = norm_cc(getv(row_vals, cc_col_idx))
                pr_val = norm_proc(getv(row_vals, pr_col_idx))

                # indicador 10
                if contrato_corto_col is not None:
                    cc_raw = norm_text(getv(row_vals, contrato_corto_col)).upper()
                    if cc_raw.startswith("S"):
                        contrato_corto_flag = "S"
                    elif cc_raw.startswith("N"):
                        contrato_corto_flag = "N"

                # indicador 52 (SI/NO)
                if indicador_52_col is not None:
                    i52_raw = norm_text(getv(row_vals, indicador_52_col)).upper()
                    if i52_raw.startswith("SI"):
                        ind52_flag = "SI"
                    elif i52_raw.startswith("NO"):
                        ind52_flag = "NO"

                rut_dv = norm_text(getv(row_vals, rut_col_idx))

                # LEY por fila (preferente LEY AFECTO)
                row_law_digits = digits_only(getv(row_vals, ley_afecto_col_idx)) if ley_afecto_col_idx else ""
                if not row_law_digits:
                    row_law_digits = law_fallback

                # filtro leyes (multi)
                if laws_filter is not None and row_law_digits and (row_law_digits not in laws_filter):
                    passes_law_row = False

                # normalizar C/J y CALIDAD
                raw_cj = getv(row_vals, cj_col_idx) if cj_col_idx else None
                raw_cal = getv(row_vals, calidad_col_idx) if calidad_col_idx else None

                cj_digits = digits_only(raw_cj)

                if not cj_digits and raw_cal is not None:
                    cj_from_cal, cal_txt = split_cj_calidad(raw_cal)
                    if cj_from_cal:
                        cj_digits = cj_from_cal
                    if cal_txt:
                        calidad_txt = cal_txt
                else:
                    if raw_cal is not None:
                        _, cal_txt = split_cj_calidad(raw_cal)
                        if cal_txt:
                            calidad_txt = cal_txt

                passes_cc_row = (selected_centros is None or cc_val in selected_centros)
                passes_pr_row = (selected_procesos is None or pr_val in selected_procesos)
                passes_filters_row = passes_cc_row and passes_pr_row and passes_law_row

                # ===== APORTE BIENESTAR (NO contrato corto S) =====
                if (
                    passes_filters_row
                    and mes_pago is not None
                    and mes_pago in selected_months
                    and rut_dv
                    and ind52_flag == "SI"
                    #   and (contrato_corto_flag != "S")  # <-- eliminado: ahora incluye S y N
                    and cj_digits
                    and row_law_digits
                ):
                    llave_bienestar = f"{cj_digits}11152{row_law_digits}"
                    bienestar_ruts_by_llave_month.setdefault(llave_bienestar, {}).setdefault(mes_pago, set()).add(rut_dv)

            # ===== escribir base (con inserción C/J si falta) + acumulación montos =====
            for c in range(1, max_col_base + 1):
                if cj_missing_in_source and cj_insert_pos == c:
                    if r_idx == base_header_row:
                        out_row.append("C/J")
                    elif r_idx >= base_data_start:
                        out_row.append(cj_digits if cj_digits else None)
                    else:
                        out_row.append(None)

                v = getv(row_vals, c)

                if r_idx >= base_data_start:
                    if cj_col_idx and c == cj_col_idx and cj_digits:
                        v = cj_digits
                    if calidad_col_idx and c == calidad_col_idx and calidad_txt is not None:
                        v = calidad_txt

           # Nota: la acumulación de montos se hace DESPUÉS, agregando por CÓDIGO (sumando columnas repetidas)


                out_row.append(v)


            # ===== acumulación montos por CÓDIGO (suma columnas repetidas) =====
            if (
                r_idx >= base_data_start
                and passes_filters_row
                and mes_pago is not None
                and mes_pago in selected_months
                and cj_digits
                and row_law_digits
            ):
                for code, cols in code_to_cols.items():
                    total_code = 0.0

                    # sumar todas las apariciones del mismo código en la fila
                    for col in cols:
                        vcol = getv(row_vals, col)
                        num = to_number(vcol)
                        if num is not None:
                            total_code += num

                    if total_code == 0.0:
                        continue

                    llave = f"{cj_digits}{code}{row_law_digits}"

                    # Acumulación general: incluye TODO
                    sums_by_llave_month.setdefault(llave, {})
                    sums_by_llave_month[llave][mes_pago] = sums_by_llave_month[llave].get(mes_pago, 0.0) + total_code

                    # Acumulación específica contrato corto='S' (solo para reglas de cálculo)
                    if contrato_corto_flag == "S":
                        sums_by_llave_month_S.setdefault(llave, {})
                        sums_by_llave_month_S[llave][mes_pago] = sums_by_llave_month_S[llave].get(mes_pago, 0.0) + total_code

                    # acumulación PF por CC (ALL / S / N) para coherencia REMU vs PF
                    if cc_val:
                        # ALL (incluye cortos y no cortos) para la mayoría de llaves
                        if selected_centros is None:
                            if cc_val not in detected_cc_set:
                                detected_cc_set.add(cc_val)
                                detected_cc_order.append(cc_val)
                        sums_by_llave_month_cc_all.setdefault(llave, {}).setdefault(mes_pago, {})
                        dcc_all = sums_by_llave_month_cc_all[llave][mes_pago]
                        dcc_all[cc_val] = dcc_all.get(cc_val, 0.0) + total_code
                    
                        # N-only (sin contratos cortos): mantiene el comportamiento histórico (v17) y sirve para llaves 'SIN cortos'
                        if contrato_corto_flag != "S":
                            sums_by_llave_month_N.setdefault(llave, {})
                            sums_by_llave_month_N[llave][mes_pago] = sums_by_llave_month_N[llave].get(mes_pago, 0.0) + total_code
                            sums_by_llave_month_cc.setdefault(llave, {}).setdefault(mes_pago, {})
                            dcc_n = sums_by_llave_month_cc[llave][mes_pago]
                            dcc_n[cc_val] = dcc_n.get(cc_val, 0.0) + total_code
                        else:
                            # S-only (solo contratos cortos): para suplencias calculadas desde componentes
                            sums_by_llave_month_cc_S.setdefault(llave, {}).setdefault(mes_pago, {})
                            dcc_s = sums_by_llave_month_cc_S[llave][mes_pago]
                            dcc_s[cc_val] = dcc_s.get(cc_val, 0.0) + total_code
            if cj_missing_in_source and cj_insert_pos == (max_col_base + 1):
                if r_idx == base_header_row:
                    out_row.append("C/J")
                elif r_idx >= base_data_start:
                    out_row.append(cj_digits if cj_digits else None)
                else:
                    out_row.append(None)

            ws_out_base.append(out_row)

        # ===== Post-proceso BIENESTAR =====
        bienestar_amount_by_llave_month: dict[str, dict[int, float]] = {}
        for llave, per_month_sets in bienestar_ruts_by_llave_month.items():
            bienestar_amount_by_llave_month[llave] = {}
            for m, rut_set in per_month_sets.items():
                bienestar_amount_by_llave_month[llave][m] = float(len(rut_set) * 15110)

        # ===== CONSOLIDACIÓN =====
        max_col_con = ws_consol.max_column or 1

        con_header_row, con_map = find_header_row_and_map(ws_consol, ["LLAVE PARA CONSOLIDADO"], search_rows=250)
        if con_header_row is None:
            raise ValueError('No pude detectar "LLAVE PARA CONSOLIDADO" en CONSOLIDACIÓN (primeras 250 filas).')

        llave_col = find_col_by_contains(con_map, "LLAVE PARA CONSOLIDADO")
        if llave_col is None:
            raise ValueError('No pude ubicar columna "LLAVE PARA CONSOLIDADO".')

        total_col = find_total_col(con_map, ley=None)
        if total_col is None:
            raise ValueError('No encontré "TOTAL LEY" en CONSOLIDACIÓN.')

        sigfe_col_con = find_col_by_contains(con_map, "CODIGO SIGFE")
        sigfe_target_by_month: dict[int, dict[str, float]] = {m: {} for m in selected_months}

        con_data_start = con_header_row + 1

        ley_col = find_col_exact(con_map, "LEY") or find_col_by_contains(con_map, "LEY")
        codigo_col = find_col_exact(con_map, "CODIGO") or find_col_by_contains(con_map, "CODIGO")
        glosa_col = find_col_exact(con_map, "GLOSA") or find_col_by_contains(con_map, "GLOSA")

        reglas_col = None
        for cand in ["REGLAS DE CALCULO", "REGLAS PARA CALCULO", "REGLAS", "CALCULO", "CHAT GPT"]:
            reglas_col = find_col_by_contains(con_map, cand)
            if reglas_col:
                break

        # detectar si ya existen columnas MES xx
        month_cols: dict[int, int] = {}
        for m in selected_months:
            col_m = find_col_exact(con_map, f"MES {m:02d}") or find_col_by_contains(con_map, f"MES {m:02d}")
            if col_m:
                month_cols[m] = col_m
        has_month_cols = (len(month_cols) == len(selected_months))

        insert_pos = total_col - 1
        month_headers = [f"MES {m:02d}" for m in selected_months]
        month_count = len(selected_months)

        # Pre-scan suplencias
        suplencias_src_by_dest: dict[str, list[str]] = {}
        if glosa_col and reglas_col:
            last_by_group: dict[tuple[str, str], list[str]] = {}
            for r in range(con_data_start, (ws_consol.max_row or 1) + 1):
                gl = norm_text(ws_consol.cell(r, glosa_col).value).upper()
                if "SUPLENCIAS" not in gl:
                    continue
                dest = digits_only(ws_consol.cell(r, llave_col).value)
                if not dest:
                    continue
                ley_digits = digits_only(ws_consol.cell(r, ley_col).value) if ley_col else ""
                group = (gl, ley_digits)
                reglas_txt = norm_text(ws_consol.cell(r, reglas_col).value)
                src_keys = re.findall(r"\b\d+\b", reglas_txt)
                if src_keys:
                    last_by_group[group] = [k.strip() for k in src_keys if k.strip()]
                    suplencias_src_by_dest[dest] = last_by_group[group]
                else:
                    if group in last_by_group:
                        suplencias_src_by_dest[dest] = last_by_group[group]

        # Escribir CONSOLIDACIÓN
        for r_idx, row_vals in enumerate(ws_consol.iter_rows(values_only=True), start=1):
            row_list = list(row_vals) if row_vals else []
            if len(row_list) < max_col_con:
                row_list.extend([None] * (max_col_con - len(row_list)))

            if r_idx == con_header_row:
                if has_month_cols:
                    ws_out_con.append(row_list)
                else:
                    ws_out_con.append(row_list[:insert_pos] + month_headers + row_list[insert_pos:])
                continue

            if r_idx < con_data_start:
                if has_month_cols:
                    ws_out_con.append(row_list)
                else:
                    ws_out_con.append(row_list[:insert_pos] + ([None] * month_count) + row_list[insert_pos:])
                continue

            llave_val = digits_only(row_list[llave_col - 1])
            codigo_txt = norm_text(row_list[codigo_col - 1]) if codigo_col else ""
            glosa_txt = norm_text(row_list[glosa_col - 1]).upper() if glosa_col else ""
            reglas_txt = norm_text(row_list[reglas_col - 1]) if reglas_col else ""
            ley_digits_row = digits_only(row_list[ley_col - 1]) if ley_col else ""

            # filtro leyes sobre CONSOLIDACIÓN
            if laws_filter is not None and ley_digits_row and (ley_digits_row not in laws_filter):
                month_vals = [0.0 for _ in selected_months]
                total = 0.0
            else:
                base_per_month = (sums_by_llave_month_N.get(llave_val, {}) if (llave_val in LLAVES_SIN_CORTOS) else sums_by_llave_month.get(llave_val, {})) if llave_val else {}

                # Regla 1: INDICADOR 54 * 0.0195 (Ley 16744 1.95)
                if "INDICADOR 54" in codigo_txt.upper() and "0.0195" in codigo_txt:
                    month_vals = [float(round(base_per_month.get(m, 0.0) * 0.0195)) for m in selected_months]
                    total = float(sum(month_vals))

                # Regla 2: Aporte Bienestar
                elif llave_val in bienestar_amount_by_llave_month:
                    per_month_amt = bienestar_amount_by_llave_month.get(llave_val, {})
                    month_vals = [float(per_month_amt.get(m, 0.0)) for m in selected_months]
                    total = float(sum(month_vals))

                # Regla 3: Reglas de cálculo (S-only solo para suplencias targets; resto ALL)
                elif reglas_txt and re.search(r"\b\d+\b", reglas_txt):
                    src_keys = [k.strip() for k in re.findall(r"\b\d+\b", reglas_txt) if k.strip()]
                    month_vals = []
                    for m in selected_months:
                        acc = 0.0
                        for k in src_keys:
                            acc += float((sums_by_llave_month_S if (llave_val in TARGET_SUPLENCIAS_SOLO_CORTOS) else sums_by_llave_month).get(k, {}).get(m, 0.0))
                        month_vals.append(acc)
                    total = float(sum(month_vals))

                # Default
                else:
                    month_vals = [float(base_per_month.get(m, 0.0)) for m in selected_months]
                    total = float(sum(month_vals))

            # acumular PF targets por SIGFE (post-reglas)
            if sigfe_col_con:
                sigfe_val = digits_only(row_list[sigfe_col_con - 1])
                if sigfe_val:
                    for i, m in enumerate(selected_months):
                        sigfe_target_by_month[m][sigfe_val] = sigfe_target_by_month[m].get(sigfe_val, 0.0) + float(month_vals[i])

            # escribir
            if has_month_cols:
                out_row = row_list[:]
                for m, col_m in month_cols.items():
                    out_row[col_m - 1] = float(month_vals[selected_months.index(m)])
                out_row[total_col - 1] = float(total)
                ws_out_con.append(out_row)
            else:
                out_row = row_list[:insert_pos] + month_vals + row_list[insert_pos:]
                out_row[insert_pos + month_count] = float(total)
                ws_out_con.append(out_row)

        # =========================
        # PF (SIGFE x CC): hojas mensuales + PF_ANUAL
        # =========================
        if generate_pf:
            ws_model = ws_pf_model

            sigfe_col = find_col_by_contains(con_map, "CODIGO SIGFE")
            if sigfe_col is None:
                raise ValueError('Marcaste "Programa Financiero", pero no encontré "CODIGO SIGFE" en CONSOLIDACIÓN.')

            # map LLAVE -> SIGFE
            llave_to_sigfe: dict[str, str] = {}
            for row_vals in ws_consol.iter_rows(min_row=con_data_start, values_only=True):
                row_list = list(row_vals) if row_vals else []
                if len(row_list) < max_col_con:
                    row_list.extend([None] * (max_col_con - len(row_list)))
                lv = digits_only(row_list[llave_col - 1])
                sv = digits_only(row_list[sigfe_col - 1])
                if lv and sv and lv not in llave_to_sigfe:
                    llave_to_sigfe[lv] = sv

            # CC columns
            if selected_centros_ordered is not None:
                cc_columns = selected_centros_ordered
            else:
                cc_columns = detected_cc_order

            cc_set = set(cc_columns)

            # sigfe_month_cc desde BASE por llave (ALL / N) + suplencias (S-only por componentes)
            sigfe_month_cc: dict[int, dict[str, dict[str, float]]] = {m: {} for m in selected_months}

            def add_dist(month: int, sigfe: str, dist: dict[str, float]):
                if month not in sigfe_month_cc:
                    return
                sm = sigfe_month_cc[month]
                sm.setdefault(sigfe, {})
                d = sm[sigfe]
                for cc, val in dist.items():
                    if cc_columns and cc not in cc_set:
                        continue
                    d[cc] = d.get(cc, 0.0) + float(val)

            # 1) Llaves "normales": usar ALL para cuadrar con CONSOL (que usa ALL),
            #    excepto LLAVES_SIN_CORTOS que usan N-only.
            for llave, per_month_all in sums_by_llave_month_cc_all.items():
                sigfe = llave_to_sigfe.get(llave, "")
                if not sigfe:
                    continue
                src_per_month = (sums_by_llave_month_cc if (llave in LLAVES_SIN_CORTOS) else sums_by_llave_month_cc_all).get(llave, {})
                for m, per_cc in src_per_month.items():
                    add_dist(m, sigfe, per_cc)

            # 2) Suplencias CALCULADAS: distribuir por CC heredando componentes S-only
            for dest_llave, src_keys in suplencias_src_by_dest.items():
                if dest_llave not in TARGET_SUPLENCIAS_SOLO_CORTOS:
                    continue
                sigfe = llave_to_sigfe.get(dest_llave, "")
                if not sigfe:
                    continue
                for m in selected_months:
                    dist_sum: dict[str, float] = {}
                    for k in src_keys:
                        per_cc = sums_by_llave_month_cc_S.get(k, {}).get(m, {})
                        for cc, val in per_cc.items():
                            dist_sum[cc] = dist_sum.get(cc, 0.0) + float(val)
                    add_dist(m, sigfe, dist_sum)

            # modelo PF
            model_header_row, model_map = find_header_row_and_map(ws_model, ["CODIGO SIGFE"], search_rows=200)
            if model_header_row is None:
                raise ValueError('No pude detectar encabezado "CODIGO SIGFE" en la hoja modelo PF.')

            col_tipo = find_col_by_contains(model_map, "TIPO")
            col_subtipo = find_col_by_contains(model_map, "SUBTIPO")
            col_sigfe_m = find_col_by_contains(model_map, "CODIGO SIGFE")
            col_den = find_col_by_contains(model_map, "DENOMINACION") or find_col_by_contains(model_map, "DENOMINACIÓN")
            if col_sigfe_m is None:
                raise ValueError('La hoja modelo no tiene columna "CODIGO SIGFE".')

            model_data_start = model_header_row + 1
            max_col_model = ws_model.max_column or 1


            # --- PF_CC Reconciliación (consistencia CC vs TOTAL) ---
            # Objetivo: evitar "descuentos" (SIN_CC negativo) cuando SUM(CC) > TOTAL.
            # No toca reglas de negocio ni el TOTAL: solo reconcilia la desagregación por CC para que sea coherente.
            pfcc_recon_log: list[dict] = []

            def _reconcile_cc_vals(
                *,
                month_label: str,
                sigfe: str,
                per_cc: dict,
                total: float,
                cc_cols: list[str],
                decimals: int = 2,
                tol: float = 0.05,
            ):
                # Construye vector en el orden de cc_cols
                vals = [float(per_cc.get(cc, 0.0) or 0.0) for cc in cc_cols]
                sum_cc = float(sum(vals))
                meta = {
                    "month": month_label,
                    "sigfe": sigfe,
                    "total": float(total),
                    "sum_cc_before": float(sum_cc),
                    "action": "none",
                    "factor": 1.0,
                    "adjustments": [],
                    "tolerance": float(tol),
                }

                # Caso trivial
                if total == 0.0 or sum_cc == 0.0:
                    sin_cc = float(total - sum_cc)
                    return vals, sin_cc, meta

                # Diferencia respecto al TOTAL (positivo => falta asignar CC, negativo => sobre-asignado)
                diff = float(total - sum_cc)

                # Si la diferencia es pequeña (redondeo), ajustar en el CC de mayor monto
                if abs(diff) <= tol:
                    # Ajuste en CC con mayor monto absoluto
                    if any(vals):
                        idx = max(range(len(vals)), key=lambda i: abs(vals[i]))
                        vals[idx] = round(vals[idx] + diff, decimals)
                        meta["action"] = "rounding_adjust"
                        meta["adjustments"].append({"cc": cc_cols[idx], "delta": float(diff)})
                        # Recalcular sum_cc tras ajuste
                        sum_cc = float(sum(vals))
                    sin_cc = 0.0
                    return vals, sin_cc, meta

                # Si falta asignar, se deja como SIN_CC (sin inventar distribución)
                if diff > 0:
                    sin_cc = float(diff)
                    return vals, sin_cc, meta

                # Si hay sobre-asignación por más que la tolerancia, NO se escala proporcionalmente
                # (eso deforma la distribución respecto a CONSOLIDADO). En su lugar, se descuenta el exceso
                # desde los CC más grandes hasta cerrar.
                excess = float(-diff)
                meta["action"] = "cap_excess"
                # Ordenar índices por monto descendente
                idxs = sorted(range(len(vals)), key=lambda i: vals[i], reverse=True)
                for idx in idxs:
                    if excess <= 0:
                        break
                    take = min(vals[idx], excess)
                    if take > 0:
                        vals[idx] = round(vals[idx] - take, decimals)
                        meta["adjustments"].append({"cc": cc_cols[idx], "delta": float(-take)})
                        excess -= take

                # Si no alcanzó (todo quedó en cero), dejamos SIN_CC en 0 y registramos exceso remanente
                if excess > 0:
                    meta["adjustments"].append({"unresolved_excess": float(excess)})

                sin_cc = 0.0
                return vals, sin_cc, meta
            def build_pf_sheet_for_month(month: int):
                ws_pf = wb_out.create_sheet(title=f"PF_MES_{month:02d}")
                headers = ["TIPO", "SUBTIPO", "CODIGO SIGFE", "DENOMINACIÓN"] + list(cc_columns) + ["SIN_CC", "TOTAL"]
                ws_pf.append(headers)

                for row_vals in ws_model.iter_rows(min_row=model_data_start, values_only=True):
                    row_list = list(row_vals) if row_vals else []
                    if len(row_list) < max_col_model:
                        row_list.extend([None] * (max_col_model - len(row_list)))

                    sigfe = digits_only(row_list[col_sigfe_m - 1])
                    tipo = row_list[col_tipo - 1] if col_tipo else None
                    subtipo = row_list[col_subtipo - 1] if col_subtipo else None
                    den = row_list[col_den - 1] if col_den else None

                    if not sigfe:
                        ws_pf.append([tipo, subtipo, None, den] + ([0.0] * len(cc_columns)) + [0.0, 0.0])
                        continue

                    per_cc = sigfe_month_cc.get(month, {}).get(sigfe, {})
                    target_total = float(sigfe_target_by_month.get(month, {}).get(sigfe, 0.0))

                    vals, sin_cc, _meta = _reconcile_cc_vals(
                        month_label=f"{month:02d}",
                        sigfe=sigfe,
                        per_cc=per_cc,
                        total=target_total,
                        cc_cols=list(cc_columns),
                    )
                    ws_pf.append([tipo, subtipo, sigfe, den] + vals + [sin_cc, target_total])

            def build_pf_sheet_anual():
                ws_pf = wb_out.create_sheet(title="PF_ANUAL")
                headers = ["TIPO", "SUBTIPO", "CODIGO SIGFE", "DENOMINACIÓN"] + list(cc_columns) + ["SIN_CC", "TOTAL"]
                ws_pf.append(headers)

                annual_cc: dict[str, dict[str, float]] = {}
                annual_target: dict[str, float] = {}

                for m in selected_months:
                    for sigfe, v in sigfe_target_by_month.get(m, {}).items():
                        annual_target[sigfe] = annual_target.get(sigfe, 0.0) + float(v)

                    sm = sigfe_month_cc.get(m, {})
                    for sigfe, per_cc in sm.items():
                        annual_cc.setdefault(sigfe, {})
                        for cc, val in per_cc.items():
                            annual_cc[sigfe][cc] = annual_cc[sigfe].get(cc, 0.0) + float(val)

                for row_vals in ws_model.iter_rows(min_row=model_data_start, values_only=True):
                    row_list = list(row_vals) if row_vals else []
                    if len(row_list) < max_col_model:
                        row_list.extend([None] * (max_col_model - len(row_list)))

                    sigfe = digits_only(row_list[col_sigfe_m - 1])
                    tipo = row_list[col_tipo - 1] if col_tipo else None
                    subtipo = row_list[col_subtipo - 1] if col_subtipo else None
                    den = row_list[col_den - 1] if col_den else None

                    if not sigfe:
                        ws_pf.append([tipo, subtipo, None, den] + ([0.0] * len(cc_columns)) + [0.0, 0.0])
                        continue

                    per_cc = annual_cc.get(sigfe, {})
                    target_total = float(annual_target.get(sigfe, 0.0))

                    vals, sin_cc, _meta = _reconcile_cc_vals(
                        month_label="ANUAL",
                        sigfe=sigfe,
                        per_cc=per_cc,
                        total=target_total,
                        cc_cols=list(cc_columns),
                    )
                    ws_pf.append([tipo, subtipo, sigfe, den] + vals + [sin_cc, target_total])

            # Hojas mensuales
            for m in selected_months:
                build_pf_sheet_for_month(m)

            # Hoja anual
            build_pf_sheet_anual()


        # --- Export log de reconciliación PF_CC (solo filas ajustadas) ---
        try:
            if pfcc_recon_log:
                import csv
                log_path = str(Path(out_path).with_suffix("")) + "_PFCC_recon_log.csv"
                with open(log_path, "w", newline="", encoding="utf-8") as f:
                    w = csv.DictWriter(
                        f,
                        fieldnames=[
                            "month", "sigfe", "total", "sum_cc_before",
                            "action", "factor", "rounding_adjust_cc", "rounding_diff"
                        ],
                    )
                    w.writeheader()
                    for r in pfcc_recon_log:
                        w.writerow(r)
        except Exception:
            # No detenemos el proceso por logging
            pass
        wb_out.save(out_path)
        return out_path

    finally:
        wb_src.close()


# =========================
# UI: Scrollable container
# =========================
class ScrollableFrame(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)

        self.canvas = tk.Canvas(self, highlightthickness=0)
        self.vsb = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.vsb.set)

        self.vsb.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)

        self.inner = ttk.Frame(self.canvas)
        self.inner_id = self.canvas.create_window((0, 0), window=self.inner, anchor="nw")

        self.inner.bind("<Configure>", self._on_inner_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)

        self.canvas.bind("<Enter>", self._bind_mousewheel)
        self.canvas.bind("<Leave>", self._unbind_mousewheel)

    def _on_inner_configure(self, event=None):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        self.canvas.itemconfigure(self.inner_id, width=event.width)

    def _bind_mousewheel(self, event=None):
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)      # Windows/Mac
        self.canvas.bind_all("<Button-4>", self._on_mousewheel_linux)  # Linux
        self.canvas.bind_all("<Button-5>", self._on_mousewheel_linux)  # Linux

    def _unbind_mousewheel(self, event=None):
        self.canvas.unbind_all("<MouseWheel>")
        self.canvas.unbind_all("<Button-4>")
        self.canvas.unbind_all("<Button-5>")

    def _on_mousewheel(self, event):
        step = -1 * int(event.delta / 120) if event.delta else 0
        if step != 0:
            self.canvas.yview_scroll(step, "units")

    def _on_mousewheel_linux(self, event):
        if event.num == 4:
            self.canvas.yview_scroll(-1, "units")
        elif event.num == 5:
            self.canvas.yview_scroll(1, "units")


# =========================
# Tkinter App
# =========================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Paso 1-7 + Programa Financiero (SIGFE x CC) -> salida con PF_MES_xx + PF_ANUAL")
        self.geometry("980x900")
        self.resizable(True, True)
        self.minsize(720, 560)

        sf = ScrollableFrame(self)
        sf.pack(fill="both", expand=True)
        root = sf.inner

        self.file_var = tk.StringVar()
        self.base_sheet_var = tk.StringVar()
        self.consol_sheet_var = tk.StringVar()
        self.pf_model_sheet_var = tk.StringVar()
        self.status_var = tk.StringVar(value="Listo.")
        self.generate_pf_var = tk.BooleanVar(value=False)

        # ===== Archivo =====
        tk.Label(root, text="Archivo Excel (.xlsx)").pack(anchor="w", padx=10, pady=(10, 0))
        ff = tk.Frame(root)
        ff.pack(fill="x", padx=10)
        tk.Entry(ff, textvariable=self.file_var).pack(side="left", fill="x", expand=True)
        tk.Button(ff, text="Buscar", command=self.select_file).pack(side="left", padx=6)

        # ===== Hojas =====
        tk.Label(root, text="Hoja BASE (origen)").pack(anchor="w", padx=10, pady=(10, 0))
        self.base_menu = tk.OptionMenu(root, self.base_sheet_var, "")
        self.base_menu.pack(fill="x", padx=10)

        tk.Label(root, text='Hoja CONSOLIDACIÓN (origen) (ej: "REPORTE REMU")').pack(anchor="w", padx=10, pady=(10, 0))
        self.consol_menu = tk.OptionMenu(root, self.consol_sheet_var, "")
        self.consol_menu.pack(fill="x", padx=10)

        tk.Label(root, text='Hoja "PROGRAMA FINANCIERO" (modelo, origen)').pack(anchor="w", padx=10, pady=(10, 0))
        self.pf_model_menu = tk.OptionMenu(root, self.pf_model_sheet_var, "")
        self.pf_model_menu.pack(fill="x", padx=10)

        # ===== PF checkbox =====
        pf_opt = tk.Frame(root)
        pf_opt.pack(fill="x", padx=10, pady=(8, 0))
        tk.Checkbutton(
            pf_opt,
            text='Generar salida "Programa Financiero" (SIGFE x Centro de Costo) (incluye PF_ANUAL)',
            variable=self.generate_pf_var
        ).pack(anchor="w")

        # ===== Leyes (multi) =====
        tk.Label(root, text="Leyes (uno/varios/todas)").pack(anchor="w", padx=10, pady=(10, 0))
        lf = tk.Frame(root)
        lf.pack(fill="x", padx=10)

        self.law_list = tk.Listbox(lf, selectmode="extended", height=3, exportselection=False)
        self.law_list.pack(side="left", fill="x", expand=True)

        for ley in ["18834", "19664", "15076"]:
            self.law_list.insert("end", ley)
        self.law_list.select_set(0, "end")

        lb = tk.Frame(lf)
        lb.pack(side="left", padx=8)
        tk.Button(lb, text="Todas", width=10, command=self.select_all_laws).pack(pady=2)
        tk.Button(lb, text="Ninguna", width=10, command=self.clear_laws).pack(pady=2)
        tk.Button(lb, text="Invertir", width=10, command=self.invert_laws).pack(pady=2)

        # ===== MES selector =====
        tk.Label(root, text="Filtro MES PAGO (uno/varios/todos)").pack(anchor="w", padx=10, pady=(12, 0))
        mf = tk.Frame(root)
        mf.pack(fill="x", padx=10)

        self.month_list = tk.Listbox(mf, selectmode="extended", height=6, exportselection=False)
        self.month_list.pack(side="left", fill="x", expand=True)
        for m in range(1, 13):
            self.month_list.insert("end", f"{m:02d}")
        self.month_list.select_set(0, "end")

        mb = tk.Frame(mf)
        mb.pack(side="left", padx=8)
        tk.Button(mb, text="Todos", width=12, command=self.select_all_months).pack(pady=2)
        tk.Button(mb, text="Ninguno", width=12, command=self.clear_months).pack(pady=2)
        tk.Button(mb, text="Invertir", width=12, command=self.invert_months).pack(pady=2)

        # ===== PROCESO selector =====
        tk.Label(root, text='Filtro "PROCESO"').pack(anchor="w", padx=10, pady=(12, 0))
        pf = tk.Frame(root)
        pf.pack(fill="both", padx=10, pady=(0, 6))

        self.proc_list = tk.Listbox(pf, selectmode="extended", height=6, exportselection=False)
        self.proc_list.pack(side="left", fill="both", expand=True)

        pb = tk.Frame(pf)
        pb.pack(side="left", padx=8, fill="y")
        tk.Button(pb, text="Todos", width=12, command=self.select_all_proc).pack(pady=2)
        tk.Button(pb, text="Ninguno", width=12, command=self.clear_proc).pack(pady=2)
        tk.Button(pb, text="Invertir", width=12, command=self.invert_proc).pack(pady=2)
        tk.Button(pb, text="Refrescar lista", width=12, command=self.refresh_proc_list).pack(pady=(10, 2))

        # ===== CENTRO DE COSTO selector =====
        tk.Label(root, text='Filtro "CENTRO DE COSTO"').pack(anchor="w", padx=10, pady=(12, 0))
        cf = tk.Frame(root)
        cf.pack(fill="both", padx=10, pady=(0, 6))

        self.cc_list = tk.Listbox(cf, selectmode="extended", height=10, exportselection=False)
        self.cc_list.pack(side="left", fill="both", expand=True)

        cb = tk.Frame(cf)
        cb.pack(side="left", padx=8, fill="y")
        tk.Button(cb, text="Todos", width=12, command=self.select_all_cc).pack(pady=2)
        tk.Button(cb, text="Ninguno", width=12, command=self.clear_cc).pack(pady=2)
        tk.Button(cb, text="Invertir", width=12, command=self.invert_cc).pack(pady=2)
        tk.Button(cb, text="Refrescar lista", width=12, command=self.refresh_cc_list).pack(pady=(10, 2))

        # ===== Botones =====
        bf = tk.Frame(root)
        bf.pack(fill="x", padx=10, pady=(12, 8))

        self.run_btn = tk.Button(
            bf,
            text="Ejecutar (Salida: BASE + CONSOL + PF opcional)",
            command=self.run_job,
            height=2
        )
        self.run_btn.pack(side="left", fill="x", expand=True)
        tk.Button(bf, text="Salir", command=self.destroy, height=2).pack(side="left", padx=8)

        # ===== Estado =====
        tk.Label(root, text="Estado").pack(anchor="w", padx=10, pady=(10, 0))
        self.pb = ttk.Progressbar(root, mode="indeterminate")
        self.pb.pack(fill="x", padx=10)
        tk.Label(root, textvariable=self.status_var).pack(anchor="w", padx=10, pady=(8, 12))

    # ===== Helpers UI Leyes =====
    def select_all_laws(self):
        self.law_list.select_set(0, "end")

    def clear_laws(self):
        self.law_list.selection_clear(0, "end")

    def invert_laws(self):
        selected = set(self.law_list.curselection())
        self.law_list.selection_clear(0, "end")
        for i in range(self.law_list.size()):
            if i not in selected:
                self.law_list.select_set(i)

    def get_selected_laws(self) -> list[str]:
        idxs = self.law_list.curselection()
        if not idxs:
            return []
        return [self.law_list.get(i) for i in idxs]

    # ===== Helpers UI Meses =====
    def select_all_months(self):
        self.month_list.select_set(0, "end")

    def clear_months(self):
        self.month_list.selection_clear(0, "end")

    def invert_months(self):
        selected = set(self.month_list.curselection())
        self.month_list.selection_clear(0, "end")
        for i in range(self.month_list.size()):
            if i not in selected:
                self.month_list.select_set(i)

    def get_selected_months(self) -> list[int]:
        idxs = self.month_list.curselection()
        if not idxs:
            return []
        out = []
        for i in idxs:
            txt = self.month_list.get(i)
            try:
                out.append(int(txt))
            except:
                pass
        return out

    # ===== Helpers UI CC =====
    def select_all_cc(self):
        self.cc_list.select_set(0, "end")

    def clear_cc(self):
        self.cc_list.selection_clear(0, "end")

    def invert_cc(self):
        selected = set(self.cc_list.curselection())
        self.cc_list.selection_clear(0, "end")
        for i in range(self.cc_list.size()):
            if i not in selected:
                self.cc_list.select_set(i)

    def get_selected_cc(self) -> list[str]:
        idxs = self.cc_list.curselection()
        if not idxs:
            return []
        return [self.cc_list.get(i) for i in idxs]

    # ===== Helpers UI Proceso =====
    def select_all_proc(self):
        self.proc_list.select_set(0, "end")

    def clear_proc(self):
        self.proc_list.selection_clear(0, "end")

    def invert_proc(self):
        selected = set(self.proc_list.curselection())
        self.proc_list.selection_clear(0, "end")
        for i in range(self.proc_list.size()):
            if i not in selected:
                self.proc_list.select_set(i)

    def get_selected_proc(self) -> list[str]:
        idxs = self.proc_list.curselection()
        if not idxs:
            return []
        return [self.proc_list.get(i) for i in idxs]

    # ===== Loaders =====
    def refresh_cc_list(self):
        src = self.file_var.get().strip()
        base_sheet = self.base_sheet_var.get().strip()
        if not src or not base_sheet:
            messagebox.showerror("Error", "Selecciona archivo y hoja BASE primero.")
            return
        self.set_busy(True, "Cargando CENTRO DE COSTO desde BASE...")
        threading.Thread(target=self._load_cc_worker, args=(src, base_sheet), daemon=True).start()

    def _load_cc_worker(self, src, base_sheet):
        try:
            centros = get_centros_costo_from_base_xlsx(src, base_sheet)

            def apply():
                self.cc_list.delete(0, "end")
                for c in centros:
                    self.cc_list.insert("end", c)
                if centros:
                    self.cc_list.select_set(0, "end")
                self.set_busy(False, f"Listo. Centros cargados: {len(centros)}")

            self.after(0, apply)
        except Exception as e:
            err_msg = str(e)
            self.after(0, lambda: self.set_busy(False, "Error."))
            self.after(0, lambda m=err_msg: messagebox.showerror("Error", m))

    def refresh_proc_list(self):
        src = self.file_var.get().strip()
        base_sheet = self.base_sheet_var.get().strip()
        if not src or not base_sheet:
            messagebox.showerror("Error", "Selecciona archivo y hoja BASE primero.")
            return
        self.set_busy(True, "Cargando PROCESO desde BASE...")
        threading.Thread(target=self._load_proc_worker, args=(src, base_sheet), daemon=True).start()

    def _load_proc_worker(self, src, base_sheet):
        try:
            procesos = get_procesos_from_base_xlsx(src, base_sheet)

            def apply():
                self.proc_list.delete(0, "end")
                for p in procesos:
                    self.proc_list.insert("end", p)
                if procesos:
                    self.proc_list.select_set(0, "end")
                self.set_busy(False, f"Listo. Procesos cargados: {len(procesos)}")

            self.after(0, apply)
        except Exception as e:
            err_msg = str(e)
            self.after(0, lambda: self.set_busy(False, "Error."))
            self.after(0, lambda m=err_msg: messagebox.showerror("Error", m))

    # ===== General =====
    def set_busy(self, busy: bool, msg: str = ""):
        if busy:
            self.run_btn.config(state="disabled")
            self.pb.start(10)
        else:
            self.pb.stop()
            self.run_btn.config(state="normal")
        if msg:
            self.status_var.set(msg)

    def select_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        self.file_var.set(path)
        self.set_busy(True, "Leyendo nombres de hojas...")
        threading.Thread(target=self._load_sheets_worker, args=(path,), daemon=True).start()

    def _load_sheets_worker(self, path: str):
        try:
            sheets = get_sheetnames_fast_xlsx(path)
            if not sheets:
                raise ValueError("No pude listar hojas (¿xlsx válido?).")

            def apply():
                m1 = self.base_menu["menu"]
                m1.delete(0, "end")
                for s in sheets:
                    m1.add_command(label=s, command=lambda name=s: self.base_sheet_var.set(name))
                self.base_sheet_var.set(sheets[0])

                m2 = self.consol_menu["menu"]
                m2.delete(0, "end")
                for s in sheets:
                    m2.add_command(label=s, command=lambda name=s: self.consol_sheet_var.set(name))
                self.consol_sheet_var.set(sheets[1] if len(sheets) > 1 else sheets[0])

                m3 = self.pf_model_menu["menu"]
                m3.delete(0, "end")
                for s in sheets:
                    m3.add_command(label=s, command=lambda name=s: self.pf_model_sheet_var.set(name))
                self.pf_model_sheet_var.set(sheets[2] if len(sheets) > 2 else sheets[0])

                self.set_busy(False, "Listo. (Carga PROCESO y CC con 'Refrescar lista')")

            self.after(0, apply)
        except Exception as e:
            err_msg = str(e)
            self.after(0, lambda: self.set_busy(False, "Error."))
            self.after(0, lambda m=err_msg: messagebox.showerror("Error", m))

    def run_job(self):
        src = self.file_var.get().strip()
        base_sheet = self.base_sheet_var.get().strip()
        consol_sheet = self.consol_sheet_var.get().strip()
        pf_model_sheet = self.pf_model_sheet_var.get().strip()
        rng = "AUTO"

        laws = self.get_selected_laws()
        if not laws:
            messagebox.showerror("Error", "Selecciona al menos 1 ley.")
            return

        months = self.get_selected_months()
        centros = self.get_selected_cc()
        procesos = self.get_selected_proc()
        generate_pf = bool(self.generate_pf_var.get())

        if not src or not base_sheet or not consol_sheet:
            messagebox.showerror("Error", "Completa: archivo, hoja base y hoja consolidación.")
            return

        if generate_pf and len(months) == 0:
            messagebox.showerror("Error", 'Para generar "Programa Financiero", debes seleccionar al menos 1 mes.')
            return

        if generate_pf and not pf_model_sheet:
            messagebox.showerror("Error", 'Selecciona la hoja "PROGRAMA FINANCIERO (modelo)".')
            return

        # Si no cargaron listas, ejecutar = "todos"
        if self.proc_list.size() == 0:
            if not messagebox.askyesno(
                "Procesos no cargados",
                'No has cargado la lista de "PROCESO".\n¿Ejecutar igual (equivale a TODOS los procesos)?'
            ):
                return
            procesos = []

        if self.cc_list.size() == 0:
            if not messagebox.askyesno(
                "Centros no cargados",
                'No has cargado la lista de "CENTRO DE COSTO".\n¿Ejecutar igual (equivale a TODOS los centros)?'
            ):
                return
            centros = []

        self.set_busy(True, "Procesando (MES + PROCESO + CENTRO DE COSTO + PF opcional)...")
        threading.Thread(
            target=self._run_worker,
            args=(src, base_sheet, consol_sheet, pf_model_sheet, rng, laws, months, centros, procesos, generate_pf),
            daemon=True
        ).start()

    def _run_worker(self, src, base_sheet, consol_sheet, pf_model_sheet, rng, laws, months, centros, procesos, generate_pf):
        try:
            law_fallback = laws[0] if laws else "18834"

            out = process_steps_1_to_4(
                src_path=src,
                base_sheet_name=base_sheet,
                consol_sheet_name=consol_sheet,
                col_range_montos=rng,
                law=law_fallback,          # fallback interno (solo si BASE no trae LEY AFECTO)
                laws_selected=laws,        # filtro multi-ley real
                months_selected=months,
                centros_selected=centros,
                procesos_selected=procesos,
                generate_pf=generate_pf,
                pf_model_sheet_name=pf_model_sheet if generate_pf else None,
            )

            self.after(0, lambda: self.set_busy(False, f"OK: {out}"))
            self.after(0, lambda: messagebox.showinfo(
                "Listo",
                f"Archivo generado:\n{out}\n\n"
                "Incluye BASE + CONSOLIDACIÓN.\n"
                + ('Incluye hojas PF_MES_xx + PF_ANUAL.\n' if generate_pf else '')
            ))
        except Exception as e:
            err_msg = str(e)
            self.after(0, lambda: self.set_busy(False, "Error."))
            self.after(0, lambda m=err_msg: messagebox.showerror("Error", m))


if __name__ == "__main__":
    App().mainloop()