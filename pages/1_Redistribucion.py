# -*- coding: utf-8 -*-
"""
Página 1 — Redistribución PRAPS vs DAP
Envuelve la lógica de app_redistribucion_mod2.py sin modificarla.
"""

import sys
import os
import streamlit as st

# Asegurar que el directorio del proyecto esté en el path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import ev_design

st.set_page_config(
    page_title="Paso 3: Redistribución de Gastos · Evidant",
    page_icon="🔄",
    layout="wide",
    initial_sidebar_state="expanded",
)


ev_design.render(
    current="redistribucion",
    page_title="Redistribución PRAPS vs DAP",
    page_sub="Paso 2 · Eliminación de reintegros y redistribución por centro de costo",
    breadcrumb="Procesamiento Financiero › Paso 2",
    icon="🔄",
)
# ── ev_design ya inyectó el CSS Spotify maestro y el page header arriba ──

# ── Importar y ejecutar la lógica del módulo original ─────────────────────────────
# Usamos importlib para evitar que el st.set_page_config() interno del módulo
# se ejecute al momento del import (el módulo tiene lógica condicional pero
# importlib garantiza que __name__ != "__main__")
import traceback
import io
import os
import tempfile
import importlib.util as _ilu

try:
    _redist_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "app_redistribucion_mod2.py",
    )
    _spec = _ilu.spec_from_file_location("app_redistribucion_mod2", _redist_path)
    _redist_mod = _ilu.module_from_spec(_spec)
    _redist_mod.__name__ = "app_redistribucion_mod2"   # ← no "__main__"
    # Registrar en sys.modules ANTES de exec_module para que @dataclass
    # pueda resolver cls.__module__ en Python 3.13
    sys.modules["app_redistribucion_mod2"] = _redist_mod
    _spec.loader.exec_module(_redist_mod)

    generate_redistribution = _redist_mod.generate_redistribution
    write_output_excel      = _redist_mod.write_output_excel

    # ── Monkey-patch: bloque de descuentos fijo (no requiere hoja Excel) ──────
    # Los datos son fijos y nunca cambian. Se inyectan directamente en el módulo
    # para que generate_redistribution los use sin leer ninguna hoja adicional.
    # La función original read_discount_block_keys queda reemplazada por esta
    # versión que devuelve las llaves pre-construidas con _join3.
    _BLOQUE_FIJO = [
        # (TIPO DE HABER,          CODIGO, DESCRIPCIÓN)
        ("HABER NO PERMANANTE", "38",  "Total renta afecta a impuesto unico"),
        ("HABER NO PERMANANTE", "44",  "Aguinaldo septiembre"),
        ("HABER NO PERMANANTE", "45",  "Aguinaldo diciembre"),
        ("HABER NO PERMANANTE", "62",  "La lb lc ld le ley 15.386 articulo 19"),
        ("HABER NO PERMANANTE", "63",  "Asig.familiar"),
        ("HABER NO PERMANANTE", "64",  "Asig. familiar duplo"),
        ("HABER NO PERMANANTE", "73",  "Bono escolaridad"),
        ("HABER NO PERMANANTE", "75",  "Adicional bono escolaridad"),
        ("HABER NO PERMANANTE", "245", "Bono ir voluntario"),
        ("HABER NO PERMANANTE", "246", "Bono adicional ir"),
        ("HABER NO PERMANANTE", "259", "Art.1 y 4 l20.646"),
        ("HABER NO PERMANANTE", "264", "Bono perman. l20743"),
        ("HABER NO PERMANANTE", "265", "Dif. he. diurnas"),
        ("HABER NO PERMANANTE", "266", "Dif. he. nocturnas"),
        ("HABER NO PERMANANTE", "267", "Bonoespecial no imp"),
        ("HABER NO PERMANANTE", "268", "Bono de vacaciones"),
        ("HABER NO PERMANANTE", "281", "Bono permanencia ir"),
        ("HABER NO PERMANANTE", "282", "Bono antiguedad ir"),
        ("HABER NO PERMANANTE", "283", "Bono trab.pesado ir"),
        ("HABER NO PERMANANTE", "295", "Otros haberes"),
        ("HABER NO PERMANANTE", "296", "Bono covid-19"),
        ("HABER PERMANANTE",    "500", "Ley 19.490 (art.1)"),
        ("HABER NO PERMANANTE", "551", "Hrs.extras diurnas"),
        ("HABER NO PERMANANTE", "552", "Hrs.extras nocturnas"),
        ("HABER PERMANANTE",    "601", "Ley 19.937 art.61."),
        ("HABER PERMANANTE",    "602", "Ley 19.937 art.64."),
        ("HABER NO PERMANANTE", "700", "Rebaja l21233 it"),
        ("HABER NO PERMANANTE", "38",  "Hrs.extras diurnas"),
        ("HABER NO PERMANANTE", "63",  "Cargas familiares"),
        ("HABER NO PERMANANTE", "64",  "Cargas fam duplo"),
        ("HABER NO PERMANANTE", "92",  "Art 19 ley 15386 (canaempu)"),
        ("HABER NO PERMANANTE", "138", "Hrs.extras nocturnas"),
        ("HABER NO PERMANANTE", "291", "1 Mes sueldos art 29  art 93"),
        ("HABER NO PERMANANTE", "700", "B.individual l19664"),
        ("HABER NO PERMANANTE", "800", "B.colectivo l19664"),
        ("HABER NO PERMANANTE", "38",  "Diferencia imponible con zona"),
        ("HABER NO PERMANANTE", "64",  "Asignacion atenciones domiciliarias"),
        ("HABER NO PERMANANTE", "88",  "Asignacion familiar"),
        ("HABER NO PERMANANTE", "291", "Un mes de sueldo art 29 l art 93"),
        ("HABER NO PERMANANTE", "296", "Art. 12 l. 20.707"),
    ]

    def _patched_read_discount_block_keys(file_like, sheet_name="BLOQUE DESCUENTO",
                                           tipo_col="TIPO DE HABER", codigo_col="CODIGO",
                                           desc_col="DESCRIPCIÓN"):
        """Devuelve las llaves del bloque fijo sin leer ninguna hoja Excel."""
        _j3 = _redist_mod._join3
        keys = [_j3(t, c, d) for t, c, d in _BLOQUE_FIJO]
        import pandas as _pd
        s = _pd.Series(keys)
        dup_keys = s[s.duplicated(keep=False)].unique().tolist()
        seen = set()
        keys_dedup = []
        for k in keys:
            if k not in seen:
                keys_dedup.append(k)
                seen.add(k)
        return keys_dedup, dup_keys

    _redist_mod.read_discount_block_keys = _patched_read_discount_block_keys
    # ─────────────────────────────────────────────────────────────────────────

    _import_ok = True
except Exception as _e:
    _import_ok = False
    _import_err = str(_e)
    _tb_import  = traceback.format_exc()

if not _import_ok:
    st.error(f"No se pudo importar app_redistribucion_mod2.py: {_import_err}")
    st.code(_tb_import)
    st.info("Asegúrate de que el archivo esté en el mismo directorio que esta suite.")
    st.stop()

# ── UI del módulo (replica de run_streamlit_app() del original) ───────────────
import pandas as pd

uploaded = st.file_uploader(
    "Sube el archivo Excel (.xlsx) — 3. REDISTRIBUCION DE GASTOS requiere solo la hoja CONSOLIDADO_REDISTRIBUIDO",
    type=["xlsx"],
    help="Sube el archivo 2. CENTRALIZACION DE GASTOS.xlsx generado en el Paso 2",
)

if not uploaded:
    st.info("⬆️  Sube el archivo generado en el Paso 2 (2. CENTRALIZACION DE GASTOS.xlsx).")
    st.stop()

# Leer hojas
try:
    xls = pd.ExcelFile(uploaded, engine="openpyxl")
    sheets = xls.sheet_names
except Exception:
    st.error("No pude leer el Excel. ¿Está corrupto o protegido con contraseña?")
    st.code(traceback.format_exc())
    st.stop()

st.info("✅ El bloque de descuentos está embebido en la suite — no necesitas subir ningún archivo adicional.")

col1, col2 = st.columns([2, 1])
with col1:
    sheet_consolidado = st.selectbox(
        "Hoja CONSOLIDADO",
        options=sheets,
        index=sheets.index("CONSOLIDADO_REDISTRIBUIDO") if "CONSOLIDADO_REDISTRIBUIDO" in sheets
               else (sheets.index("CONSOLIDADO") if "CONSOLIDADO" in sheets else 0),
    )
with col2:
    header_rows = st.number_input(
        "Filas encabezado compuesto",
        min_value=1, max_value=6, value=3, step=1,
        help="Normalmente 3 filas en el CONSOLIDADO",
    )

# Bloque fijo — no se necesita selección de hoja
sheet_bloque = "BLOQUE DESCUENTO"  # valor dummy, el patch no lo usa

st.divider()

if st.button("⚡  GENERAR REDISTRIBUCIÓN", type="primary", use_container_width=True):
    try:
        with st.status("⚡ Generando Redistribución...", expanded=True) as _red_status:
          _pbar_red = st.progress(0.1, text="Procesando redistribución...")
          with st.spinner("Procesando redistribución..."):
            uploaded.seek(0)
            consolidado_out, resumen, estado, audit = generate_redistribution(
                input_path_or_buffer=uploaded,
                sheet_consolidado=sheet_consolidado,
                sheet_bloque=sheet_bloque,
                header_rows=int(header_rows),
            )

            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp_path = tmp.name

            # ── Ejecutar motor original sin modificaciones ────────────────────
            write_output_excel(tmp_path, consolidado_out, resumen, estado, audit)

            # ── Post-proceso: restaurar encabezado compuesto 3 filas en
            #    CONSOLIDADO_REDISTRIBUIDO (el motor escribe 1 fila plana con
            #    nombres "p1 | p2 | p3"). No tocamos ninguna otra hoja. ────────
            from openpyxl import load_workbook as _lw_redist
            _wb_r = _lw_redist(tmp_path)
            _ws_r = _wb_r["CONSOLIDADO_REDISTRIBUIDO"]

            # Leer fila 1 (encabezados planos "p1 | p2 | p3")
            flat_headers = [cell.value for cell in _ws_r[1]]

            # Descomponer cada columna en sus 3 partes
            def _split3(col_name):
                if col_name is None:
                    return ("", "", "")
                parts = str(col_name).split(" | ", 2)
                while len(parts) < 3:
                    parts.append("")
                return tuple(parts[:3])

            h1_r = [_split3(h)[0] for h in flat_headers]
            h2_r = [_split3(h)[1] for h in flat_headers]
            h3_r = [_split3(h)[2] for h in flat_headers]

            # Leer todas las filas de datos (fila 2 en adelante)
            data_rows_r = []
            for row in _ws_r.iter_rows(min_row=2, values_only=True):
                data_rows_r.append(list(row))

            # Reconstruir hoja: limpiar y reescribir
            _ws_r.delete_rows(1, _ws_r.max_row)
            _ws_r.append(h1_r)
            _ws_r.append(h2_r)
            _ws_r.append(h3_r)
            for row in data_rows_r:
                _ws_r.append(row)

            _wb_r.save(tmp_path)
            _wb_r.close()
            # ─────────────────────────────────────────────────────────────────

            with open(tmp_path, "rb") as f:
                out_bytes = f.read()

            try:
                os.remove(tmp_path)
            except Exception:
                pass

        st.success("✅ Procesamiento completado exitosamente.")
        st.download_button(
            "📥  Descargar 3. REDISTRIBUCION DE GASTOS.xlsx",
            data=out_bytes,
            file_name="3. REDISTRIBUCION DE GASTOS.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        st.subheader("📊 Resumen de Auditoría")
        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("Pares reintegro borrados",    audit.reintegro_pairs_deleted)
        c2.metric("Reintegros no pareados",       audit.reintegro_unpaired_rows)
        c3.metric("Neto negativo (casos)",        audit.neto_negative_rows)
        c4.metric("Conceptos bloque no hallados", len(audit.missing_discount_columns))
        c5.metric("Llaves duplicadas en bloque",  len(audit.duplicate_discount_keys_in_block))

        st.subheader("Vista previa: Resumen por (Centro de Costo, Programa)")
        st.markdown(ev_design.ev_table_html(resumen.head(50)), unsafe_allow_html=True)

        if audit.missing_discount_columns:
            st.warning("⚠️ Hay conceptos del BLOQUE DESCUENTO que NO se encontraron en el CONSOLIDADO.")
            st.markdown(
                ev_design.ev_table_html(
                    pd.DataFrame({"MISSING_DISCOUNT_KEYS": audit.missing_discount_columns}).head(100)
                ),
                unsafe_allow_html=True,
            )

    except Exception:
        st.error("❌ La ejecución falló.")
        st.code(traceback.format_exc())
