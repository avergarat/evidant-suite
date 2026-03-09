# -*- coding: utf-8 -*-
"""
Página 4 — Consolidación Reportes Remu
Motor: CONSOLIDACIO_N_REPORTE_REMU.py (importado sin ejecutar Tkinter)

Post-procesamiento:
  1. Elimina columnas __BLANK__ del consolidado
  2. Renombra la columna CENTRO DE COSTO (fila-3) → SIRH
  3. Inserta dos columnas nuevas inmediatamente después de DV:
       • CENTRO DE COSTO  → número de resolución (homologado desde SIRH)
       • PROGRAMA         → nombre oficial del programa (homologado desde SIRH)
  La homologación usa _HOMOLOG_PROG_LIST definido en este módulo.
  Archivo de salida: "1. CONSOLIDADO_<nombre_original>.xlsx"
"""

import sys
import os
import io
import re
import traceback
import importlib.util
import tempfile
import unicodedata

import streamlit as st
import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import ev_design

# Repositorio RR.HH.
try:
    from repositorio.db import upsert_desde_consolidado, generar_id_contrato, H3_TO_CAMPO
    _repo_ok = True
except Exception:
    _repo_ok = False




st.set_page_config(
    page_title="Paso 1: Consolidación Reportería SIRH · Evidant",
    page_icon="📂",
    layout="wide",
    initial_sidebar_state="expanded",
)

ev_design.render(
    current="consolidacion",
    page_title="Consolidación de Reportes",
    page_sub="Paso 1 · Consolidación multi-hoja con encabezado triple y repositorio RR.HH.",
    breadcrumb="Procesamiento Financiero › Paso 1",
    icon="📂",
)
# ═══════════════════════════════════════════════════════════════════════════════
# TABLA DE HOMOLOGACIÓN DE PROGRAMAS
# Estructura: (RES, NOMBRE_OFICIAL, [variantes_informales...])
# ═══════════════════════════════════════════════════════════════════════════════
_HOMOLOG_PROG_LIST = [
    # 53 - APOYO GESTIÓN (base)
    ("53", "APOYO A LA GESTION EN LOS ESTABLECIMIENTOS DEPENDIENTES DE LOS SERVICIOS DE SALUD", [
        "apoyo a la gestion en el nivel primario de salud en los establecimientos dependientes de los servicios de salud",
        "apoyo a la gestion en el nivel primario de salud en los establecimientos dependientes de los servicios de salud apertura",
        "apoyo a la gestion en el nivel primario de salud en los establecimientos dependientes de los servicios de salud continuidad de la atencion",
        "apoyo a la gestion en el nivel primario de salud en los establecimientos dependientes de los servicios de salud controles",
        "apoyo a la gestion en el nivel primario de salud en los establecimientos dependientes de los servicios de salud ecografias",
        "apoyo a la gestion en el nivel primario de salud en los establecimientos dependientes de los servicios de salud ecografias extra",
        "apoyo a la gestion en el nivel primario de salud en los establecimientos dependientes de los servicios de salud extension",
        "apoyo a la gestion en el nivel primario de salud en los establecimientos dependientes de los servicios de salud oirs",
        "apoyo a la gestion en el nivel primario de salud en los establecimientos dependientes de los servicios de salud podologia",
        "apoyo a la gestion en el nivel primario de salud en los establecimientos dependientes de los servicios de salud podologia extra",
        "apoyo a la gestion", "apoyo acciones de salud", "apoyo gestion aps",
        "consultorio dr jose ahues", "consultorio dra ana juricic", "consultorio enf sofia pincheira",
        "consultorio los nogales", "consultorio maipu", "consultorio norman voullieme",
        "oficina apoyo clinico", "vacunatorio",
        "5000001 apoyo acciones de salud", "5000002 apoyo gestion aps",
        "28 consultorio dr. jose ahues", "25 consultorio dra. ana juricic",
        "24 consultorio enf. sofia pincheira", "29 consultorio los nogales",
        "30 consultorio maipu", "23 consultorio norman voullieme",
        "340450 oficina apoyo clinico", "2581780 vacunatorio",
        "467010 apoyo a la gestion",
    ]),
    # 53 - APOYO GESTIÓN ACOMPAÑAMIENTO
    ("53", "APOYO A LA GESTION EN LOS ESTABLECIMIENTOS DEPENDIENTES DE LOS SERVICIOS DE SALUD - ACOMPAÑAMIENTO", [
        "apoyo a la gestion en el nivel primario de salud en los establecimientos dependientes de los servicios de salud acompanamiento",
        "5000051 apoyo acciones de salud  acompanamiento",
        "5000051 apoyo acciones de salud  acompanamiento",
    ]),
    # 53 - Campaña vacunación
    ("53", "APOYO A LA GESTION EN EL NIVEL PRIMARIO DE SALUD EN LOS ESTABLECIMIENTOS DEPENDIENTES DE LOS SERVICIOS DE SALUD/CAMPAÑA DE VACUNACION", [
        "apoyo a la gestion en el nivel primario de salud en los establecimientos dependientes de los servicios de salud campana de vacunacion extra",
    ]),
    # 179 - PESPI
    ("179", "PROGRAMA ESPECIAL DE SALUD Y PUEBLOS INDIGENAS (PESPI)", [
        "programa especial de salud y pueblos indigenas pespi",
    ]),
    # 180 - MAIS
    ("180", "MODELO DE ATENCION INTEGRAL DE SALUD FAMILIAR Y COMUNITARIA EN ATENCION PRIMARIA (MAIS)", [
        "programa mais", "500064 programa mais",
    ]),
    # 211 - ELIGE VIDA SANA
    ("211", "ELIGE VIDA SANA", [
        "elige vida sana", "elige vida sana consultas", "elige vida sana rescates",
        "elige vida sana talleres", "vida sana factores de riesgo enferm",
        "5000027 vida sana factores de riesgo enferm",
    ]),
    # 212 - MIGRANTES
    ("212", "ACCESO A LA ATENCION DE SALUD A PERSONAS MIGRANTES", [
        "acceso a la atencion de salud a personas migrantes dupla sanitaria",
        "acceso a la atencion de salud a personas migrantes dupla sanitaria apertura",
        "acceso a la atencion de salud a personas migrantes dupla sanitaria divap",
        "acceso a la atencion de salud a personas migrantes dupla sanitaria extension",
    ]),
    # 214 - CECOSF
    ("214", "CENTRO COMUNITARIOS DE SALUD FAMILIAR (CECOSF)", [
        "centros comunitarios de salud familiar  cecosf  gestor",
        "centros comunitarios de salud familiar  cecosf",
        "cecosf", "centros comunitarios de salud familiar cecosf",
        "centros comunitarios de salud familiar cecosf gestor",
        "5000009 cecosf",
    ]),
    # 216 - FOFAR
    ("216", "FONDO DE FARMACIA PARA ENFERMEDADES CRONICAS NO TRANSMISIBLES EN APS (FOFAR)", [
        "fofar  fondo farmacia", "fofar", "fondo de farmacia",
        "5000013 fofar  fondo farmacia",
    ]),
    # 217 - BUCAL sub-líneas
    ("217", "ESTRATEGIA DE SALUD BUCAL/ENSEÑANZA MEDIA", [
        "estrategias de salud bucal odontologico integral ensenanza media",
    ]),
    ("217", "ESTRATEGIA DE SALUD BUCAL/GES", [
        "estrategias de salud bucal ges odontologico", "ges odontologico",
        "5000015 ges odontologico",
    ]),
    ("217", "ESTRATEGIA DE SALUD BUCAL/GES (PROTESIS)", [
        "estrategias de salud bucal ges odontologico protesis", "ges odontologico protesis",
        "estrategias de salud bucal ges odontologico  protesis",
    ]),
    ("217", "ESTRATEGIA DE SALUD BUCAL/HOMBRES DE ESCASOS RECURSOS", [
        "estrategias de salud bucal odontologico integral hombres de escasos recursos",
    ]),
    ("217", "ESTRATEGIA DE SALUD BUCAL/HOMBRES DE ESCASOS RECURSOS (PROTESIS)", [
        "estrategias de salud bucal hombres de escasos recursos protesis",
        "estrategias de salud bucal odontologico integral hombres de escasos recursos protesis",
        "estrategias de salud bucal odontologico integral hombres de escasos recursos  protesis",
    ]),
    ("217", "ESTRATEGIA DE SALUD BUCAL/MAS SONRISAS", [
        "estrategias de salud bucal odontologico integral mas sonrisas para chile",
    ]),
    ("217", "ESTRATEGIA DE SALUD BUCAL/MAS SONRISAS (PROTESIS)", [
        "estrategias de salud bucal odontologico integral mas sonrisas para chile protesis",
        "estrategias de salud bucal odontologico integral mas sonrisas para chile  protesis",
    ]),
    ("217", "ESTRATEGIA DE SALUD BUCAL/MEJORAMIENTO", [
        "estrategias de salud bucal mejoramiento del acceso a la atencion odontologica",
        "estrategias de salud bucal mejoramiento del acceso a la atencion odontologica jornada",
        "estrategias de salud bucal mejoramiento del acceso a la atencion odontologica  jornada",
    ]),
    # 226 - CONTINUIDAD CUIDADOS
    ("226", "CONTINUIDAD DE CUIDADOS PREVENTIVOS Y DE TRATAMIENTO EN ATENCION PRIMARIA DE SALUD", [
        "continuidad de cuidados preventivos y de tratamiento en aps control ginecologico",
        "continuidad de cuidados preventivos y de tratamiento en atencion primaria de salud control ginecologico",
    ]),
    # 228 - MAMAV
    ("228", "MAS ADULTOS MAYORES AUTOVALENTES", [
        "mas adultos mayores autovalentes  mamav",
        "mas adultos mayores autovalentes", "mas adultos mayores autovalentes mamav",
        "5000017 mas adultos mayores autovalentes",
    ]),
    # 229 - REHABILITACION INTEGRAL
    ("229", "REHABILITACION INTEGRAL EN APS", [
        "rehabilitacion integral", "5000022 rehabilitacion integral",
    ]),
    # 231 - SALUD RESPIRATORIA
    ("231", "SALUD RESPIRATORIA", [
        "era", "ira", "salud respiratoria era", "salud respiratoria ira", "salud respiratoria",
        "5000011 era", "5000016 ira",
    ]),
    ("231", "SALUD RESPIRATORIA/CAMPAÑA INVIERNO", [
        "salud respiratoria campana de invierno",
    ]),
    ("231", "SALUD RESPIRATORIA/IRA EN SAPU", [
        "salud respiratoria ira en sapu",
    ]),
    ("231", "SALUD RESPIRATORIA/KINE EXTENSION", [
        "salud respiratoria kinesiologo extension",
    ]),
    ("231", "SALUD RESPIRATORIA/REFUERZO SAPU-SAR", [
        "salud respiratoria refuerzo sapu sar", "sapu", "5000025 sapu",
    ]),
    # 232 - ESPACIOS AMIGABLES
    ("232", "ESPACIOS AMIGABLES PARA ADOLESCENTES", [
        "espacios amigables para adolescentes",
        "espacios amigables para adolescentes extension",
    ]),
    # 233 - SAR Pincheira
    ("233", "SERVICIO DE ATENCION PRIMARIA DE URGENCIA DE ALTA RESOLUCION (SAR)", [
        "sar  pincheira", "sar pincheira",
        "servicios de atencion primaria de urgencia de alta resolucion  sar",
        "5000044 sar  pincheira",
    ]),
    # 303 - SAR genérico / estrategias urgencia
    ("303", "ESTRATEGIAS DE INTERVENCION DE URGENCIA EN ATENCION PRIMARIA DE SALUD", [
        "estrategias de intervencion de urgencia en atencion primaria",
        "servicios de atencion primaria de urgencia de alta resolucion sar",
        "servicios de atencion primaria de urgencia de alta resolucion sar diurno",
        "servicios de atencion primaria de urgencia de alta resolucion sar turnos",
        "servicios de atencion primaria de urgencia de alta resolucion  sar  turnos",
        "servicios de atencion primaria de urgencia de alta resolucion  sar  diurno",
        "servicios de atencion primaria de urgencia de alta resolucion sar turnos extra",
        "servicios de atencion primaria de urgencia de alta resolucion  sar  turnos extra",
    ]),
    # 230 - IMAGENES DIAGNOSTICAS
    ("230", "IMAGENES DIAGNOSTICAS EN ATENCION PRIMARIA", [
        "imagenes diagnosticas en ap", "5000063 imagenes diagnosticas en ap",
    ]),
    # 306 - RESOLUTIVIDAD
    ("306", "RESOLUTIVIDAD EN ATENCION PRIMARIA", [
        "resolutividad en aps",
        "resolutividad en atencion primaria cirugia menor",
        "resolutividad en atencion primaria uapo",
        "resolutividad en atencion primaria uapo extra",
        "resolutividad en la atencion primaria de salud cirugia menor",
        "resolutividad en la atencion primaria de salud uapo",
        "5000023 resolutividad en aps",
    ]),
    # 308 - SALUD MENTAL APS
    ("308", "SALUD MENTAL EN LA ATENCION PRIMARIA DE SALUD", [
        "salud mental", "salud mental en la atencion primaria de salud",
        "5000024 salud mental", "2554 salud mental",
    ]),
    # 374 - DEPENDENCIA SEVERA
    ("374", "ATENCION DOMICILIARIA PARA PERSONAS CON DEPENDENCIA SEVERA Y CUIDADORES", [
        "at dom ic discap severa", "at. domic. discap. severa",
        "5000005 at. domic. discap. severa",
    ]),
    # 396 - PADB
    ("396", "PROGRAMA DE APOYO AL DESARROLLO BIOPSICOSOCIAL EN LA RED ASISTENCIAL", [
        "apoyo al desarrollo biopsicosocial en la red asistencial chile crece contigo",
        "chile crece contigo  padb", "chile crece contigo padb",
        "programa de apoyo al desarrollo biopsicosocial en la red asistencial chile crece contigo  padb",
        "5000010 chile crece contigo  padb",
    ]),
    # 810 - PASMI
    ("810", "PROGRAMA DE APOYO A LA SALUD MENTAL INFANTIL (PASMI)", [
        "programa de apoyo a la salud mental infantil  pasmi",
        "programa de apoyo a la salud mental infantil pasmi",
        "5000058 apoyo a la salud mental infantil  pasmi",
    ]),
    # S/R - DAP ADMINISTRACIÓN
    ("S/R", "D.A.P. (ADMINISTRACION)", [
        "asesoria juridica", "d.a.p.  administracion", "depto. auditoria", "depto. finanzas",
        "direccion", "direccion dap", "remuneraciones", "servicio de bienestar",
        "sin programa - apoyo a la gestion abastecimiento y logistica",
        "sin programa - apoyo a la gestion abastecimiento y logistica - apertura",
        "sin programa - apoyo a la gestion abastecimiento y logistica - extension",
        "sin programa - apoyo a la gestion alcoholemia",
        "sin programa - apoyo a la gestion alcoholemia - extra",
        "sin programa - apoyo a la gestion operaciones",
        "subdireccion administrativa", "subdireccion medica", "sudireccion medica",
        "unidad salud trabajador",
        "5000043 d.a.p.  administracion", "4 subdireccion administrativa",
        "124014 asesoria juridica", "2 sudireccion medica", "292 subdireccion medica",
        "392 subdireccion medica", "1 direccion", "111000000 direccion dap",
        "114 depto. auditoria", "418 depto. finanzas", "422608 remuneraciones",
        "422658 servicio de bienestar", "422734 unidad salud trabajador",
        "12401400 sin descripcion",
    ]),
    # S/R - LISTA DE ESPERA
    ("S/R", "LISTA DE ESPERA", [
        "lista de espera extra",
    ]),
    # S/R - INTERMEDIO MEDICO QUIRURGICO
    ("S/R", "INTERMEDIO MEDICO QUIRURGICO", [
        "288270 intermedio medico quirurgico", "intermedio medico quirurgico",
    ]),
]


# ═══════════════════════════════════════════════════════════════════════════════
# Motor de homologación
# ═══════════════════════════════════════════════════════════════════════════════
def _normalize_for_lookup(text: object) -> str:
    """
    Normaliza texto para búsqueda:
    - Minúsculas
    - Sin tildes / diacríticos
    - Guiones bajos y paréntesis → espacio (para manejar códigos SIRH como 5000001_APOYO...)
    - Sin dobles espacios
    """
    if text is None or (isinstance(text, float) and pd.isna(text)):
        return ""
    s = str(text).strip()
    # Quitar tildes
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    # Minúsculas
    s = s.lower()
    # Normalizar separadores de código SIRH
    s = s.replace("_", " ").replace("(", " ").replace(")", " ")
    # Colapsar espacios
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _build_lookup() -> dict:
    """Construye dict normalizado → (res, nombre_oficial)"""
    lookup: dict = {}
    for res, nombre_oficial, variantes in _HOMOLOG_PROG_LIST:
        # El nombre oficial también es clave
        k_oficial = _normalize_for_lookup(nombre_oficial)
        if k_oficial and k_oficial not in lookup:
            lookup[k_oficial] = (res, nombre_oficial)
        for v in variantes:
            k = _normalize_for_lookup(v)
            if k and k not in lookup:
                lookup[k] = (res, nombre_oficial)
    return lookup


_HOMOLOG_LOOKUP = _build_lookup()


def homologar_programa(sirh_value: object) -> tuple:
    """
    Dado el valor de la columna SIRH (ej: '5000001_APOYO ACCIONES DE SALUD'),
    retorna (res, nombre_oficial).
    Estrategia:
      1) Buscar con valor completo normalizado
      2) Buscar con el texto DESPUÉS del primer '_'
      3) Sin match: ('', valor_original)
    """
    raw = ""
    if sirh_value is not None and not (isinstance(sirh_value, float) and pd.isna(sirh_value)):
        raw = str(sirh_value).strip()

    if not raw:
        return ("", "")

    # 1) Búsqueda completa (incluye código numérico)
    key_full = _normalize_for_lookup(raw)
    if key_full in _HOMOLOG_LOOKUP:
        return _HOMOLOG_LOOKUP[key_full]

    # 2) Solo texto después del primer '_'
    if "_" in raw:
        after = raw.split("_", 1)[1].strip()
        key_after = _normalize_for_lookup(after)
        if key_after in _HOMOLOG_LOOKUP:
            return _HOMOLOG_LOOKUP[key_after]

    # 3) Sin match
    return ("", raw)


# ═══════════════════════════════════════════════════════════════════════════════
# Post-procesamiento — opera sobre listas de encabezado + filas de datos
# Retorna: (h1_out, h2_out, h3_out, data_rows_out, sirh_col_idx)
# ═══════════════════════════════════════════════════════════════════════════════
def post_process_headers_and_data(
    h1: list, h2: list, h3: list, data_rows: list
) -> tuple:
    """
    Trabaja directamente sobre las listas de encabezado y datos para:
      1. Detectar e ignorar columnas __BLANK__ (no las incluye en la salida)
      2. Detectar columna CENTRO DE COSTO (fila3) → en salida se llama SIRH
         pero conserva h1 y h2 originales de esa columna
      3. Detectar posición de DV (fila3) → insertar después:
           - col nueva: h1="", h2="", h3="CENTRO DE COSTO" → valor = RES homologado
           - col nueva: h1="", h2="", h3="PROGRAMA"        → valor = nombre oficial
    Retorna: (h1_out, h2_out, h3_out, data_out, n_sin_match, sin_match_vals)
    """
    n_src = max(len(h1), len(h2), len(h3))

    def _nc(lst, i):
        v = lst[i] if i < len(lst) else None
        return "" if v is None else str(v).strip()

    def _triplete_ends(idx, target):
        return _nc(h3, idx).upper() == target.upper()

    # ── Identificar columnas BLANK a excluir ─────────────────────────────────
    # Una columna es BLANK si:
    #   - las 3 filas están vacías, O
    #   - CUALQUIERA de las 3 filas contiene "__BLANK__" (texto literal del motor)
    blank_indices = set()
    for i in range(n_src):
        a, b, c = _nc(h1, i), _nc(h2, i), _nc(h3, i)
        all_empty = (a == "" and b == "" and c == "")
        any_blank_tag = ("__BLANK__" in a or "__BLANK__" in b or "__BLANK__" in c)
        if all_empty or any_blank_tag:
            blank_indices.add(i)

    # Índice de CENTRO DE COSTO (en h3) → pasará a llamarse SIRH en h3
    cc_idx = None
    for i in range(n_src):
        if i not in blank_indices and "CENTRO DE COSTO" in _nc(h3, i).upper():
            cc_idx = i
            break

    # Índice de RUT-DV (en h3) → insertar CENTRO DE COSTO y PROGRAMA justo después
    # Busca por: "RUT-DV", "RUT DV", o que contenga "RUT" y "DV"
    dv_idx = None
    for i in range(n_src):
        if i in blank_indices:
            continue
        h3v = _nc(h3, i).upper()
        if "RUT-DV" in h3v or "RUT DV" in h3v or h3v == "RUT-DV":
            dv_idx = i
            break
    if dv_idx is None:
        # fallback: buscar la primera columna que contenga "DV" en h3
        for i in range(n_src):
            if i not in blank_indices and _nc(h3, i).upper().strip() == "DV":
                dv_idx = i
                break

    # ── Construir columnas de salida (sin blanks) en orden original ───────────
    # Cada elemento: (src_col_idx, h1_val, h2_val, h3_val)
    out_cols = []
    for i in range(n_src):
        if i in blank_indices:
            continue
        a = _nc(h1, i)
        b = _nc(h2, i)
        c = _nc(h3, i)
        # Si es la columna CC → h3 pasa a "SIRH", h1 y h2 se conservan
        if i == cc_idx:
            c = "SIRH"
        out_cols.append((i, a, b, c))

    # ── Posición de inserción (después de DV en out_cols) ─────────────────────
    insert_after = len(out_cols)  # default: al final
    if dv_idx is not None:
        for j, (src_i, _, _, _) in enumerate(out_cols):
            if src_i == dv_idx:
                insert_after = j + 1
                break

    # ── Procesar filas de datos: homologar SIRH ───────────────────────────────
    # Primero obtener los valores de la columna SIRH (ex-CC)
    sirh_values = []
    if cc_idx is not None:
        for row in data_rows:
            v = row[cc_idx] if cc_idx < len(row) else None
            sirh_values.append(v)
    else:
        sirh_values = [None] * len(data_rows)

    # Homologar
    homolog_results = [homologar_programa(v) for v in sirh_values]
    res_vals     = [r[0] for r in homolog_results]
    prog_vals    = [r[1] for r in homolog_results]

    # Estadísticas sin match
    sin_match_set = set()
    n_sin_match = 0
    for v, (res, _) in zip(sirh_values, homolog_results):
        if res == "":
            n_sin_match += 1
            if v is not None and str(v).strip():
                sin_match_set.add(str(v).strip())

    # ── Insertar columnas nuevas en out_cols ──────────────────────────────────
    # Marcador especial src_col_idx = -1 → CENTRO DE COSTO, -2 → PROGRAMA
    col_cc_new   = (-1, "", "", "CENTRO DE COSTO")
    col_prog_new = (-2, "", "", "PROGRAMA")
    out_cols.insert(insert_after,     col_cc_new)
    out_cols.insert(insert_after + 1, col_prog_new)

    # ── Reconstruir h1_out, h2_out, h3_out ───────────────────────────────────
    h1_out, h2_out, h3_out = [], [], []
    for (src_i, a, b, c) in out_cols:
        h1_out.append(a)
        h2_out.append(b)
        h3_out.append(c)

    # ── Reconstruir filas de datos ────────────────────────────────────────────
    data_out = []
    for row_i, row in enumerate(data_rows):
        new_row = []
        for (src_i, _, _, _) in out_cols:
            if src_i == -1:
                new_row.append(res_vals[row_i])
            elif src_i == -2:
                new_row.append(prog_vals[row_i])
            else:
                new_row.append(row[src_i] if src_i < len(row) else None)
        data_out.append(new_row)

    return h1_out, h2_out, h3_out, data_out, n_sin_match, sorted(sin_match_set)


# ── ev_design ya inyectó el CSS Spotify maestro arriba, nada más necesario ──

# ── Sidebar ────────────────────────────────────────────────────────────────────

# ── Importar motor original ───────────────────────────────────────────────────
try:
    _cons_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "CONSOLIDACIO_N_REPORTE_REMU.py",
    )
    _spec = importlib.util.spec_from_file_location("consolidacion_remu", _cons_path)
    _cons_mod = importlib.util.module_from_spec(_spec)
    _cons_mod.__name__ = "consolidacion_remu"
    sys.modules["consolidacion_remu"] = _cons_mod
    _spec.loader.exec_module(_cons_mod)
    consolidate_by_header_triplet = _cons_mod.consolidate_by_header_triplet
    _import_ok = True
except Exception as _e:
    _import_ok = False
    _import_err = str(_e)
    _tb_import  = traceback.format_exc()

if not _import_ok:
    st.error(f"No se pudo importar CONSOLIDACIO_N_REPORTE_REMU.py: {_import_err}")
    st.code(_tb_import)
    st.stop()

# ── Info card ─────────────────────────────────────────────────────────────────
st.markdown("""
<div style="background:rgba(0,87,255,0.07);border:1px solid rgba(0,87,255,0.2);
            border-radius:10px;padding:0.9rem 1.2rem;margin-bottom:1.2rem;">
  <strong style="color:#90caf9;">PRIMER PASO: Consolidación Reportería Leyes SIRH</strong>
  <span style="color:#6b8caf;font-size:0.85rem;">
    Consolida los reportes de las leyes <strong style="color:#90caf9;">19.664 · 18.834 · 15.076</strong> por triplete de encabezados,
    elimina columnas en blanco, renombra <code style="color:#00e5ff;">CENTRO DE COSTO</code> → <code style="color:#00e5ff;">SIRH</code>,
    inserta <code style="color:#00e5ff;">CENTRO DE COSTO</code> (N° resolución) y <code style="color:#00e5ff;">PROGRAMA</code> después de <code style="color:#00e5ff;">RUT-DV</code>,
    y agrega las 3 hojas de la <strong style="color:#90caf9;">Plantilla CONS REMU_PF</strong> al archivo de salida
    listo para el <strong style="color:#90caf9;">Paso 2: Procesamiento PF CASA</strong>.
  </span>
</div>
""", unsafe_allow_html=True)

# ── Upload ────────────────────────────────────────────────────────────────────
col_up1, col_up2 = st.columns(2)

with col_up1:
    st.markdown("""
    <div style="font-size:0.72rem;font-weight:700;color:#6b8caf;letter-spacing:1.5px;margin-bottom:0.4rem;">
      ARCHIVO DE REPORTERÍA (LEYES SIRH)
    </div>
    """, unsafe_allow_html=True)
    uploaded = st.file_uploader(
        "Reportes de Leyes 19.664 · 18.834 · 15.076",
        type=["xlsx"],
        key="cons_uploader",
        help="Excel con las hojas de los reportes de remuneraciones por ley",
    )

with col_up2:
    st.markdown("""
    <div style="font-size:0.72rem;font-weight:700;color:#6b8caf;letter-spacing:1.5px;margin-bottom:0.4rem;">
      PLANTILLA CONS REMU_PF (3 hojas)
    </div>
    """, unsafe_allow_html=True)
    uploaded_plantilla = st.file_uploader(
        "PLANTILLA CONS REMU_PF.xlsx",
        type=["xlsx"],
        key="plantilla_uploader",
        help="Sube la plantilla con las 3 hojas que se agregarán al archivo de salida para el Paso 2",
    )

if not uploaded:
    st.info("⬆️  Sube el archivo de reportería para ver las hojas disponibles.")
    st.stop()


@st.cache_data(show_spinner=False)
def save_uploaded(file_bytes: bytes) -> str:
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, prefix="cons_src_")
    tmp.write(file_bytes)
    tmp.close()
    return tmp.name


file_bytes = uploaded.read()
src_path   = save_uploaded(file_bytes)

# Leer nombres de hojas
try:
    from openpyxl import load_workbook as _lw
    _wb        = _lw(src_path, read_only=True, data_only=False, keep_links=False)
    all_sheets = _wb.sheetnames
    _wb.close()
except Exception as e:
    st.error(f"No pude leer las hojas: {e}")
    st.stop()

st.success(f"✅ Archivo cargado · {len(all_sheets)} hojas detectadas")

# ── Selección de hojas ────────────────────────────────────────────────────────
st.markdown("### Selección y orden de hojas")

col1, col2 = st.columns([2, 1])
with col1:
    selected_sheets = st.multiselect(
        "Hojas a consolidar (el orden importa — la primera define el orden base de columnas)",
        options=all_sheets,
        default=all_sheets[:3] if len(all_sheets) >= 3 else all_sheets,
    )
with col2:
    add_source = st.checkbox(
        "Agregar columna ORIGEN_HOJA",
        value=True,
        help="Añade una columna al inicio indicando de qué hoja proviene cada fila",
    )

if selected_sheets:
    order_html = " → ".join(
        f'<code style="color:#00e5ff;font-size:0.75rem;">{s}</code>'
        for s in selected_sheets
    )
    st.markdown(
        f'<div style="font-size:0.78rem;color:#6b8caf;margin:0.3rem 0 0.6rem 0;">'
        f'Orden: {order_html}</div>',
        unsafe_allow_html=True,
    )

st.divider()

# ── Ejecutar ──────────────────────────────────────────────────────────────────
if st.button("⚡  CONSOLIDAR Y HOMOLOGAR", type="primary", use_container_width=True):
    if not selected_sheets:
        st.warning("⚠️  Selecciona al menos una hoja.")
        st.stop()

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, prefix="cons_raw_") as _tmp:
        raw_out_path = _tmp.name

    try:
        with st.status("⚡ Consolidando y Homologando...", expanded=True) as _cons_status:
          _pbar_c = st.progress(0, text="Iniciando...")
        # ── Paso 1: Motor original → archivo crudo ────────────────────────────
          _pbar_c.progress(1/7, text="Paso 1/7 — Consolidando hojas...")
          with st.spinner("Paso 1/7 — Consolidando hojas..."):
            _cons_result = consolidate_by_header_triplet(
                file_bytes,
                selected_sheets,
                add_source_col=add_source,
            )
            with open(raw_out_path, "wb") as _f:
                _f.write(_cons_result.read())

        # ── Paso 2: Leer consolidado crudo (encabezado 3 filas + datos) ─────────
          _pbar_c.progress(2/7, text="Paso 2/7 — Leyendo y homologando...")
          with st.spinner("Paso 2/7 — Leyendo y homologando..."):
            from openpyxl import load_workbook as _lw2

            _wb2 = _lw2(raw_out_path, read_only=True, data_only=True)
            _ws2 = _wb2["CONSOLIDADO"]

            rows_iter = _ws2.iter_rows(values_only=True)
            h1 = list(next(rows_iter, []))
            h2 = list(next(rows_iter, []))
            h3 = list(next(rows_iter, []))

            # Leer todas las filas de datos; padear a max_col si es necesario
            n_hdr = max(len(h1), len(h2), len(h3))
            raw_data_rows = []
            for row in rows_iter:
                r = list(row)
                if len(r) < n_hdr:
                    r.extend([None] * (n_hdr - len(r)))
                elif len(r) > n_hdr:
                    r = r[:n_hdr]
                raw_data_rows.append(r)
            _wb2.close()

            # Filtrar filas completamente vacías
            raw_data_rows = [r for r in raw_data_rows if any(v is not None and str(v).strip() != "" for v in r)]

        # ── Paso 3: Post-procesamiento sobre encabezados y datos ─────────────
          _pbar_c.progress(3/7, text="Paso 3/7 — Insertando columnas CC y Programa...")
          with st.spinner("Paso 3/7 — Insertando columnas CC y Programa..."):
            h1_out, h2_out, h3_out, data_out, n_sin_match, sin_match_vals = \
                post_process_headers_and_data(h1, h2, h3, raw_data_rows)

        # ── Paso 3b: Agregar columna ID_CONTRATO ─────────────────────────────
          _pbar_c.progress(4/7, text="Paso 4/7 — Generando ID contrato...")
          with st.spinner("Paso 4/7 — Generando ID contrato..."):
            from repositorio.db import generar_id_contrato, upsert_desde_consolidado

            # Insertar ID_CONTRATO como primera columna
            h1_out  = [""] + h1_out
            h2_out  = [""] + h2_out
            h3_out  = ["ID_CONTRATO"] + h3_out

            # Detectar índices de campos clave para generar el ID
            # (los índices ya incluyen el +1 del desplazamiento)
            def _idx(name):
                try:
                    return h3_out.index(name)
                except ValueError:
                    return None

            idx_rut  = _idx("RUT-DV")
            idx_corr = _idx("CORR")
            idx_ley  = _idx("LEY AFECTO")
            idx_hora = _idx("HORAS / GRADOS")
            idx_cc   = _idx("CENTRO DE COSTO")

            data_con_id = []
            for row in data_out:
                def _v(idx):
                    if idx is None or idx - 1 >= len(row):
                        return ""
                    v = row[idx - 1]  # -1 porque row aún no tiene el ID
                    return str(v).strip() if v is not None else ""

                id_c = generar_id_contrato(
                    _v(idx_rut), _v(idx_corr), _v(idx_ley),
                    _v(idx_hora), _v(idx_cc)
                )
                data_con_id.append([id_c] + list(row))

            data_out = data_con_id

        # ── Paso 3c: Upsert en repositorio SQLite ────────────────────────────
        # Pasamos h3/data SIN la columna ID_CONTRATO (pos 0) que agregamos
        # nosotros — el motor de la BD genera su propio ID internamente.
          _pbar_c.progress(5/7, text="Paso 5/7 — Actualizando repositorio...")
          with st.spinner("Paso 5/7 — Actualizando repositorio..."):
            try:
                repo_stats = upsert_desde_consolidado(
                    h3_out[1:],
                    [row[1:] for row in data_out],
                    uploaded.name,
                )
                _repo_ok = True
            except Exception as _re:
                _repo_ok = False
                _repo_err = str(_re)

        # ── Paso 4: Exportar con encabezado compuesto + hojas de plantilla ────
          _pbar_c.progress(6/7, text="Paso 6/7 — Exportando archivo final...")
          with st.spinner("Paso 6/7 — Exportando archivo final..."):
            out_filename = f"1. CONSOLIDADO_{uploaded.name}"
            buf = io.BytesIO()

            from openpyxl import Workbook as _WB
            from openpyxl import load_workbook as _lw_plant

            # Crear libro de salida (modo normal para poder copiar hojas)
            _wb_out = _WB()
            # Eliminar hoja por defecto
            _wb_out.remove(_wb_out.active)

            # ── Hoja 1: CONSOLIDADO (encabezado compuesto + datos) ────────────
            _ws_out = _wb_out.create_sheet("CONSOLIDADO")
            _ws_out.append(h1_out)
            _ws_out.append(h2_out)
            _ws_out.append(h3_out)
            for row in data_out:
                _ws_out.append(row)

            # ── Hojas 2-4: copiar desde plantilla si fue subida ───────────────
            plantilla_hojas_ok = []
            plantilla_hojas_err = []

            if uploaded_plantilla is not None:
                try:
                    _plant_bytes = uploaded_plantilla.read()
                    _plant_tmp = tempfile.NamedTemporaryFile(
                        suffix=".xlsx", delete=False, prefix="plantilla_"
                    )
                    _plant_tmp.write(_plant_bytes)
                    _plant_tmp.close()

                    _wb_plant = _lw_plant(
                        _plant_tmp.name, read_only=True,
                        data_only=True, keep_links=False
                    )

                    for sheet_name in _wb_plant.sheetnames:
                        _ws_src = _wb_plant[sheet_name]
                        _ws_dst = _wb_out.create_sheet(sheet_name)
                        for row in _ws_src.iter_rows(values_only=True):
                            _ws_dst.append(list(row))
                        plantilla_hojas_ok.append(sheet_name)

                    _wb_plant.close()
                    try:
                        os.remove(_plant_tmp.name)
                    except Exception:
                        pass

                except Exception as _pe:
                    plantilla_hojas_err.append(str(_pe))

            _wb_out.save(buf)

        try:
            os.remove(raw_out_path)
        except Exception:
            pass

        # ── Métricas ──────────────────────────────────────────────────────────
        total_filas = len(data_out)
        sin_match   = n_sin_match
        con_match   = total_filas - sin_match
        prog_uniq   = len(set(r[h3_out.index("PROGRAMA")] for r in data_out if "PROGRAMA" in h3_out)) if "PROGRAMA" in h3_out else 0

        n_cols_out = len(h3_out)
        n_hojas_total = 1 + len(plantilla_hojas_ok)
        st.success(f"✅ Proceso completado · {total_filas:,} filas · {n_cols_out} columnas · {n_hojas_total} hojas en archivo de salida")

        # Feedback repositorio
        if _repo_ok:
            st.info(
                f"🗄️ Repositorio actualizado — "
                f"**{repo_stats['nuevos']}** nuevos · "
                f"**{repo_stats['actualizados']}** actualizados · "
                f"**{repo_stats['sin_cambio']}** sin cambio · "
                f"**{repo_stats['inactivos']}** marcados inactivos"
            )
        else:
            st.warning(f"⚠️ Repositorio no actualizado: {_repo_err}")

        # Feedback plantilla
        if uploaded_plantilla is None:
            st.warning("⚠️  No se subió la Plantilla CONS REMU_PF — el archivo de salida solo contiene la hoja CONSOLIDADO. Para el Paso 2 (PF CASA) se requieren las 3 hojas de la plantilla.")
        elif plantilla_hojas_err:
            st.error(f"❌ Error al leer la plantilla: {plantilla_hojas_err[0]}")
        else:
            st.info(f"📋 Hojas de plantilla agregadas: {' · '.join(plantilla_hojas_ok)}")

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Total filas",           f"{total_filas:,}")
        c2.metric("Hojas consolidadas",    len(selected_sheets))
        c3.metric("Filas homologadas",     f"{con_match:,}")
        c4.metric("Sin match (revisar)",   f"{sin_match:,}",
                  delta=f"-{sin_match}" if sin_match > 0 else "✓",
                  delta_color="inverse" if sin_match > 0 else "normal")

        st.download_button(
            f"📥  Descargar  {out_filename}",
            data=buf.getvalue(),
            file_name=out_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        # ── Advertencia valores sin match ─────────────────────────────────────
        if sin_match > 0 and sin_match_vals:
            with st.expander(
                f"⚠️  {sin_match} filas sin homologación "
                f"({len(sin_match_vals)} valores SIRH únicos no reconocidos)"
            ):
                st.markdown(
                    ev_design.ev_table_html(pd.DataFrame({"SIRH_SIN_MATCH": sin_match_vals})),
                    unsafe_allow_html=True,
                )
                st.caption(
                    "Agrega estos valores a _HOMOLOG_PROG_LIST en el módulo "
                    "4_Consolidacion_Remu.py para que sean reconocidos en futuras ejecuciones."
                )

        # ── Vista previa (solo columnas clave, evita duplicados en h3) ─────────
        st.subheader("Vista previa — primeras 40 filas (columnas clave)")
        priority_h3 = ["ORIGEN_HOJA", "CENTRO DE COSTO", "PROGRAMA", "SIRH",
                        "RUT-DV", "PROCESO", "FOLIO", "AÑO PAGO", "MES PAGO",
                        "NOMBRE", "ESTAB", "CALIDAD JURIDICA", "PLANTA", "UNIDAD"]
        # Buscar índice de cada columna prioritaria (primera ocurrencia)
        preview_indices = []  # lista de (col_idx, display_name)
        used_idx = set()
        for p in priority_h3:
            for idx, name in enumerate(h3_out):
                if name == p and idx not in used_idx:
                    preview_indices.append((idx, name))
                    used_idx.add(idx)
                    break
        if preview_indices:
            preview_data = [[row[i] if i < len(row) else None
                             for i, _ in preview_indices]
                            for row in data_out[:40]]
            preview_col_names = [name for _, name in preview_indices]
            df_preview = pd.DataFrame(preview_data, columns=preview_col_names)
            st.markdown(ev_design.ev_table_html(df_preview), unsafe_allow_html=True)
        else:
            st.info("Vista previa no disponible — descarga el archivo para ver el resultado completo.")

    except Exception:
        st.error("❌ El proceso falló.")
        st.code(traceback.format_exc())
        try:
            os.remove(raw_out_path)
        except Exception:
            pass
